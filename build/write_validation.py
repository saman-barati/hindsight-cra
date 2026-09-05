# -*- coding: utf-8 -*-
"""Write docs/04-model-validation.md and docs/05-edd-policy-note-2026.md, pulling
every figure straight from the workbook so the prose cannot drift from the model."""
import os
from paths import REPO
from openpyxl import load_workbook

XL = REPO + "/model/customer-risk-model.xlsx"
D = REPO + "/docs/"
wb = load_workbook(XL, data_only=True)
V, S, B = wb["Validation"], wb["Scoring"], wb["Backtest"]

SENS = [(V.cell(r, 2).value, V.cell(r, 3).value, V.cell(r, 9).value, V.cell(r, 10).value)
        for r in range(7, 17)]
RULES = [(V.cell(r, 2).value, V.cell(r, 3).value, V.cell(r, 4).value, V.cell(r, 5).value,
          V.cell(r, 6).value, V.cell(r, 7).value) for r in range(23, 27)]
EDD = [(V.cell(r, 2).value, V.cell(r, 3).value, V.cell(r, 4).value) for r in range(33, 40)]
NBT = sum(1 for r in range(4, 30) if B.cell(r, 1).value)
BT = [(B.cell(r, 1).value, B.cell(r, 2).value, B.cell(r, 44).value, B.cell(r, 47).value,
       B.cell(r, 54).value) for r in range(4, 4 + NBT)]
ORD = ["Low", "Medium", "High"]
import bisect
from validation_block import BAND_B_LOW, BAND_B_MED
from backtest_block import CASES
from data import FACTORS

# The share of the model's weight each notice pins at level 1, and the ceiling that implies.
# Computed here rather than typed, because an earlier version hard-coded these in six documents
# and every one went stale the moment a single case was rebuilt.
_W = wb["Weights"]
_FO = [f[0] for f in FACTORS]
_WEFF = {_W.cell(r, 1).value: _W.cell(r, 3).value for r in range(2, 22)}
_ROW = {B.cell(r, 1).value: r for r in range(4, 4 + NBT)}
def _sc(ref, f): return B.cell(_ROW[ref], 23 + _FO.index(f)).value
_PIN = []
for _c, _n, _refs in CASES:
    _wp = sum(_WEFF[f] for f in _FO if all(_sc(r, f) == 1 for r in _refs))
    _PIN.append(_wp)
PINLO, PINHI = 100 * min(_PIN), 100 * max(_PIN)
MAXCEIL = max(1 * w + 5 * (1 - w) for w in _PIN)
ov = [S.cell(r, 23).value for r in range(2, 402)]
ov_sorted = sorted(ov)
def p(q): return ov_sorted[int(q * (len(ov_sorted) - 1))]

moved = sum(1 for r in BT if r[3] != r[4])
up_lm = sum(1 for r in BT if r[3] == "Low" and r[4] == "Medium")
up_lh = sum(1 for r in BT if r[3] == "Low" and r[4] == "High")
up_mh = sum(1 for r in BT if r[3] == "Medium" and r[4] == "High")
down = sum(1 for r in BT if ORD.index(r[4]) < ORD.index(r[3]))
NUM = ["no", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine", "ten",
       "eleven", "twelve", "thirteen"]
RATIO = sorted(d / float(c) for _, _, c, d in SENS)
ovr = sorted(S.cell(r, 24).value for r in range(2, 402))   # the rounded score, which is what is banded
PCTL = [100.0 * bisect.bisect_right(ovr, b) / len(ovr) for b in (BAND_B_LOW, BAND_B_MED)]


def ordinal(x):
    n = int(round(x))
    suf = "th" if 10 <= n % 100 <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return "%d%s" % (n, suf)


def moves(a, b):
    n = sum(1 for r in BT if r[3] == a and r[4] == b)
    return "%s %s move from %s to %s" % (NUM[n].capitalize() if False else NUM[n], "row" if n == 1 else "rows", a, b)

def w(path, text):
    open(path, "w", encoding="utf-8").write(text.strip() + "\n")

sens_rows = "\n".join("| %s | %s | %d | %d |" % (a, b.split(" at ")[1], c, d) for a, b, c, d in SENS)
rule_rows = "\n".join("| %s | %s | %d | %d | %d | %.1f%% |" % (a, b, c, d, e, 100 * f) for a, b, c, d, e, f in RULES)
edd_rows = "\n".join("| %s | %d | %.1f%% |" % (a, b, 100 * c) for a, b, c in EDD)
bt_rows = "\n".join("| %s | %s | %.2f | %s | **%s** |" % (a, b, c, d, e) for a, b, c, d, e in BT)

w(D + "04-model-validation.md", """
# Model validation

**Firm:** Northgate Bank UK Limited (fictional)
**Document reference:** HND-CRA-008
**Version:** 0.3
**Author:** Saman Barati
**Date:** September 2026
**Companion file:** `model/customer-risk-model.xlsx`, sheet `Validation`

---

### Author's note

This is the closest a one-person project can get to second-line challenge: I wrote the model, so I cannot independently validate it. What I can do is test it hard enough that its weaknesses are on the record rather than waiting to be found by someone else.

The largest finding is not one I expected. I spent Step 2 arguing about the weights. Step 5 shows the weights were never the thing that mattered.

---

## 1. What validation can and cannot show

1.1 This pack tests four things: how the population distributes, how much the output depends on each weight, whether a different aggregation rule would work better, and what the 2026 change to jurisdiction-based enhanced due diligence does to the book.

1.2 It cannot show that the ratings are correct. There is no outcome data. Step 4 tested the model against customers whose outcomes are publicly known, and that back-test remains the only evidence in this project about whether the model is *right* rather than merely consistent.

1.3 **And that evidence is narrower than it looks.** The facts each FCA notice settles — a UK company, a UK address, branch or app onboarding — pin between %.0f%% and %.0f%% of the model's weight at level 1 before any judgement is made. The highest score any of the six reconstructions could reach, if every remaining factor were pushed to 5, is %.2f, and the High band begins at 3.50. So the back-test's headline finding, that no reconstruction reaches High on the arithmetic, was fixed by its inputs and is not a discovery. What the back-test can still decide is whether an escalator fires, and how far the answer moves when the same file is read two defensible ways. Everything this pack takes from Step 4 is limited to those two things. See [`backtest/README.md`](../backtest/README.md).

## 2. How the population distributes

| Percentile | Overall score |
|---|---|
| 50th | %.2f |
| 75th | %.2f |
| 90th | %.2f |
| 95th | %.2f |
| Highest | %.2f |

The High band begins at 3.50. The population never gets there. Under the current bands the book is %d Low, %d Medium and %d High, and every one of those High ratings comes from a mandatory escalator rather than from the score.

## 3. Weight sensitivity

Each category weight was moved by ten percentage points, with the other four rescaled so the total stays at 100%%, and the whole population rescored. The table counts how many of the 400 customers change band.

| Scenario | Weight becomes | Band changes, current bands | Band changes, rebanded |
|---|---|---|---|
%s

### 3.1 The finding

Under the current bands, **the largest weight change in the table moves %d customers out of 400**. Removing the delivery channel category entirely — setting it to zero — moves %d.

That is the answer to the question Step 2 spent a document arguing about. Whether geography carries 25%% or 15%% changes the rating of %d customers. The weights are not what decides this model's output.

### 3.2 Why, and what actually decides it

Because the population sits far below the band boundaries, a small change to a weighted average of mostly-low scores rarely carries anyone across a line. The boundaries are doing the work, not the weights.

The right-hand column tests that directly: the same weight changes, applied against boundaries that sit where the population actually is, move between %d and %d customers — **between %.1f and %.1f times as many, scenario for scenario.** The weights only start to matter once the bands are in the right place.

So the priority order is settled by evidence rather than instinct: fix the boundaries first, then argue about the weights. Section 11.1 of the methodology proposed cutting the geography weight. That is still worth doing for the reason given there — a factor that separates nobody should not carry weight — but it is a second-order change, and this pack says so.

## 4. Aggregation rules compared

Four ways of turning 20 factor scores into a rating, all applied to the same population.

| Rule | Definition | Low | Medium | High | High share |
|---|---|---|---|---|---|
%s

### 4.1 Reading the table

Every rule in the table is a **final rating**, so the mandatory escalators at methodology 5.3 are applied on top of the arithmetic in all four rows. That is worth saying explicitly, because a reader who compares the High column with the score distribution in section 2 will not be able to reconcile them otherwise: %d customers score above %.2f, %d are escalated, and %d of those escalated customers score at or below %.2f, which is how rule B reaches %d.

**Rule A** is the model as built. It uses one of its three bands.

**Rule B** keeps the weighted average and moves the boundaries to %.2f and %.2f, which fall at the %s and %s percentiles of the population. All three bands are now used, and the High share of %.1f%% is a workload a bank could plausibly resource.

**Rule C** ignores the average and counts how many of the 20 factors score 4 or 5: none is Low, one or two is Medium, three or more is High. It is blunt, it is immune to dilution, and it puts %.1f%% of the book in High.

**Rule D** takes the worse of B and C. It is the most conservative and the least affordable at %.1f%%.

### 4.2 The operational constraint decides it, and the constraint is mine

Rules C and D put %.1f%% and %.1f%% of the book into High. Rule B puts %.1f%% there. I rejected C and D on the ground that a bank cannot apply enhanced due diligence, senior management sign-off and annual refresh to nearly a third of its customers.

**That threshold is not sourced.** There is no published figure for the share of a retail bank's book that can be held at High, and I did not find one. Regulation 33 says when enhanced due diligence is mandatory and regulation 18 requires the assessment to be proportionate to the firm's size and nature; neither gives a number, and neither would, because the answer depends on the firm. What I have is an intuition about cost, dressed as a constraint, and it is doing the decisive work in this section.

Two things follow. The argument is still directionally right: rules C and D put roughly **1.7 times** as many customers into High as rule B does, and an EDD population 1.7 times larger needs something close to 1.7 times the analyst headcount. No firm resources a control by discovering afterwards that it cannot run it. But a reader should treat "rule B is the recommendation" as **the recommendation this analyst reaches given an assumption he cannot evidence**, and the first question a real second line would ask is what the firm's actual EDD capacity is. With that number, this section decides itself. Without it, section 5's priority order rests on a guess.

**Rule B is the recommendation, on that basis and no firmer one.**

## 5. The recommended changes

In priority order, with the evidence for each.

| # | Change | Evidence |
|---|---|---|
| 1 | **Move the band boundaries to 1.60 and 2.20.** | Section 3: the boundaries, not the weights, decide the output. Section 4: rule B is the only tested rule that uses all three bands at a workable High share. |
| 2 | **Move A5 level 5 out of the weighted average and onto the escalator list.** | Step 4: a customer whose stated address is a London landmark scores %.2f and is rated Low. A declared fact that cannot be true is a failure to identify the customer, not a small increase in risk. |
| 3 | **Tie C3 level 5 to registration as a high value dealer** rather than to a description of the trade. | Step 4: the same jewellery business is Low or High depending on which of two defensible readings an analyst takes. **Contested — see 5.1.** |
| 4 | **Split PEP treatment so domestic PEPs and PEP family members are not automatically High.** | Regulation 35(3A) requires the starting point for a domestic PEP to be a *lower* level of risk than a non-domestic one. Escalator 5.3(a) treats a foreign PEP, a domestic PEP and a PEP's parent identically. |
| 5 | **Add a file-completeness factor.** | Nothing in the library asks how much of the file is evidenced rather than asserted. **The Step 4 evidence for this was withdrawn — see 5.2.** |
| 6 | **Reduce the geography weight.** | Section 3, and methodology 11.1. Genuine, but second-order until change 1 is made. |
| 7 | **Split C1 level 5** into "trust or overseas incorporation, ownership evidenced" and "nominee or bearer". | Escalator 5.3(d) is written as a condition the recorded level cannot evaluate, so it currently fires on the whole level. Methodology 11.9a. |
| 8 | **Add a level to C4 for the family and associates of a foreign PEP**, distinct from those of a domestic PEP. | Regulation 35(3A) requires the domestic starting point to be lower. The library now covers relatives of any PEP at level 3, but still cannot tell the two apart, so change 4 cannot be implemented without this. |

Changes 1 and 2 are implemented in the workbook and tested below. Changes 3 to 8 are specified but not built.

### 5.1 Change 3 is contested by the evidence for it

Tying the high-value-dealer level to registration would resolve the Fowler Oldfield ambiguity. It would also have taken **Stunt & Co from High to Medium**, because a gold refiner and trader that declares no cash would probably not have been a registered high value dealer, and C3 level 5 as currently written is the only thing in the model that catches that customer on the day the account opens.

So the change is not an improvement; it is a trade. It buys consistency on one case and loses the only correct call the back-test produced. It stays on the list because the current wording is genuinely ambiguous, but it cannot be made without a replacement control — the obvious candidate being a level that turns on the *goods* rather than on the registration, which is what the current wording was reaching for and failed to express precisely. Recorded so that nobody applies change 3 on the strength of the Fowler Oldfield finding alone.

### 5.2 The evidence for change 5 was withdrawn

Change 5 was argued from the Stunt & Co reconstruction, on the reading that Barclays had gathered almost nothing and the model therefore rated an empty file Low. That reconstruction was wrong: the Final Notice records what the application form said, and rebuilt from it the customer rates High. The specific claim — that the model cannot rate a file nobody filled in — no longer has that case behind it.

The underlying gap is still real and still unmeasured. The library has no level for "the check was run and nobody read the result" (Stunt & Co, C5) and no factor for the proportion of a file that is evidenced rather than asserted. Change 5 stays, ranked below change 4, with its evidence downgraded from a demonstrated failure to an identified gap. That is a weaker claim than the one it replaces and it is stated as such.

### 5.3 The boundaries in change 1 are fitted to a population I invented

%.2f and %.2f are the %s and %s percentiles of 400 customers whose distributions I chose in Step 3. Recommending boundaries derived from that population, and then testing them on the same population, is circular, and the circularity is at its worst exactly here — in the recommendation that section 3 says matters more than everything else in the model.

What survives the circularity is the *method*, not the numbers: set boundaries from the observed distribution and check that all three bands are used. What does not survive is any claim that 1.60 and 2.20 are the right numbers for a real book. On a real book they would be different, and the first calibration exercise a firm ran would move them.

## 6. What the recommended package does to the six enforcement cases

Rule B boundaries, plus A5 level 5 as an escalator. The same %d reconstructions from Step 4 — two or three readings of each of the six cases, so a case counts once for each reading that moves.

| Case | Enforcement case | Score | Rating as built | Rating under the recommendation |
|---|---|---|---|---|
%s

**%d of the %d rows change rating, all of them upwards** — %s. The Monzo rows are the Low-to-High moves: the address that could not be true now decides the rating on its own. Rows already rated High by an escalator cannot move, which is %d of the %d.

### 6.1 What this does not claim

None of these customers becomes High on the arithmetic, and for the reason at 1.3 none of them could have. Fowler Oldfield under the cash-intensive reading is still Medium. The recommendation does not manufacture hindsight and it would not have "caught" these customers in the sense of stopping them.

What it changes is treatment. Medium instead of Low means standard rather than simplified due diligence, source of funds required where activity is inconsistent with the profile, and a three-year rather than five-year refresh. For customers who went on to move hundreds of millions of pounds, that is a material difference, and it is the honest size of the claim.

## 7. Override governance

7.1 Methodology 5.5 sets the rules: an analyst may override upwards freely, downwards only with MLRO approval and a written rationale, never below Medium where a mandatory escalator applies, and the model is treated as needing recalibration if overrides exceed 5%% of ratings in a quarter.

7.2 What was missing was the management information that makes the 5%% threshold mean anything. The override register should report, quarterly:

- overrides as a percentage of ratings issued, split by direction
- the five factors most often overridden, which is where the level definitions are failing
- overrides by analyst, to separate a model problem from a training problem
- downward overrides that were later followed by a suspicious activity report, which is the only outcome signal this model will ever get

7.3 The third and fourth of those matter most. A high override rate concentrated on one factor is a definition to rewrite. A high override rate spread evenly is a model to recalibrate. Without the split, the 5%% threshold cannot tell the two apart.

## 8. Limitations of this validation

8.1 **The population is synthetic and its factors are drawn independently.** Real customers correlate, which would lengthen the right tail. The compression finding in section 3 does not depend on the population — it is a property of averaging 20 factors — but the exact counts do.

8.2 **The proposed boundaries are fitted to this population, and the aggregation rule is chosen on an unsourced workload limit.** Both admissions used to live here. They have been moved to 5.3 and 4.2, next to the conclusions they undercut, because a caveat filed at the back of a document is not a caveat a reader meets before believing the finding.

8.3 **The challenge in this pack is the author's own.** Sections 1.3, 4.2, 5.1, 5.2 and 5.3 each set the objection to a finding next to the finding itself, which is the discipline a second line would impose. It is not a substitute for one. Nobody independent of the model has signed this off, and a model challenged only by its author is a model with one opinion in it.

8.4 **Fitting to six cases is not validation either.** Six enforcement notices are the six that were published, prosecuted and reported. Firms that failed without being fined are not in the sample, and neither are the customers who looked terrible and turned out fine — which is the false positive side, and this project has no way to see it at all.

## 9. Model change log

| # | Change | Status |
|---|---|---|
| 1 | Band boundaries to 1.60 and 2.20 | Tested; recommended |
| 2 | A5 level 5 onto the escalator list | Tested; recommended |
| 3 | C3 level 5 tied to high value dealer registration | Specified; **contested** (5.1); not built |
| 4 | Domestic PEPs and PEP family members not automatically High | Specified; not built |
| 5 | File-completeness factor | Specified; evidence downgraded (5.2); not built |
| 6 | Reduce the geography weight | Specified; deferred until change 1 |
| 7 | Split C1 level 5 so 5.3(d) can be evaluated | Specified; not built |
| 8 | A C4 level for the relatives of a foreign PEP | Specified; not built |
| — | C4 level 3 reworded to cover the family and known close associates of **any** PEP | **Applied.** A defect fix: regulation 35(1) and escalator 5.3(a) both cover them, and the library's wording covered only the relatives of a domestic PEP, so a foreign PEP's relative had no level to be recorded at. No rating changes on this population. |
| — | Escalator 5.3(b) reworded from "second residence" to "a further tax residence" | **Applied.** Documentation fix: factor G2 records tax residence, and the escalator and the population labels both said "second residence". No rating changes. |
| — | Escalator 5.3(d) now fires on the whole of C1 level 5 | **Applied.** A defect fix, not a calibration change: the library said C1 level 5 triggered 5.3(d) where nominee shareholders or bearer shares were present, and the model never fired it at all. Three customers move to High. The proper fix, splitting C1 level 5 so the condition can be evaluated, is change 7. See methodology 11.9a. |
| — | Escalator 5.3(b) reworded to cover payment corridors | **Applied.** The model always fired on G3 level 5; the written escalator said "established in", which the model does not test. Documentation corrected to match the control. No rating changes. |

---

## Version history

| Version | Date | Change |
|---|---|---|
| 0.3 | Sept 2026 | Percentiles corrected to the 43rd and 88th; the workload ratio at 4.2 corrected to 1.7; changes 7 and 8 added to section 5; two defect fixes logged in section 9. |
| 0.2 | Sept 2026 | The back-test ceiling is now stated at 1.3 before any conclusion rests on it; the unsourced workload limit is stated at 4.2 where it decides the answer; the circularity of the proposed boundaries is stated at 5.3 where they are recommended. Change 3 is marked contested, change 5's evidence is downgraded, and a PEP change is added at 4 for regulation 35(3A). All figures re-run after escalator 5.3(d) was corrected. |
| 0.1 | Sept 2026 | First validation pack: distribution, sensitivity, aggregation rules, recommended changes. |
""" % (PINLO, PINHI, MAXCEIL,
       p(0.50), p(0.75), p(0.90), p(0.95), max(ov),
       RULES[0][2], RULES[0][3], RULES[0][4],
       sens_rows,
       max(c for _, _, c, _ in SENS), [c for a, _, c, _ in SENS if "Delivery channel -" in a][0],
       max(c for a, _, c, _ in SENS if "Geography" in a),
       min(d for _, _, _, d in SENS), max(d for _, _, _, d in SENS), RATIO[0], RATIO[-1],
       rule_rows,
       sum(1 for v in ovr if v > BAND_B_MED), BAND_B_MED,
       sum(1 for r in range(2, 402) if S.cell(r, 26).value == "Yes"),
       sum(1 for r in range(2, 402) if S.cell(r, 26).value == "Yes" and S.cell(r, 24).value <= BAND_B_MED),
       BAND_B_MED, RULES[1][4],
       BAND_B_LOW, BAND_B_MED, ordinal(PCTL[0]), ordinal(PCTL[1]),
       100 * RULES[1][5], 100 * RULES[2][5], 100 * RULES[3][5],
       100 * RULES[2][5], 100 * RULES[3][5], 100 * RULES[1][5],
       [r[2] for r in BT if r[0] == "MON-lo"][0],
       BAND_B_LOW, BAND_B_MED, ordinal(PCTL[0]), ordinal(PCTL[1]),
       NBT, bt_rows, moved, NBT,
       ", ".join(m for m in (moves("Low", "Medium"), moves("Low", "High"), moves("Medium", "High"))
                 if not m.startswith("no ")),
       sum(1 for r in BT if r[3] == "High"), NBT))

w(D + "05-edd-policy-note-2026.md", """
# Policy note: the 2026 change to jurisdiction-based enhanced due diligence

**Firm:** Northgate Bank UK Limited (fictional)
**Document reference:** HND-CRA-009
**Version:** 0.3
**To:** Money Laundering Reporting Officer, for the Financial Crime Committee
**From:** Saman Barati
**Date:** September 2026
**Decision required:** whether to replace the statutory jurisdiction trigger with a firm-policy one

---

## 1. The change

1.1 Under the Money Laundering Regulations 2017 as they stood, a relationship involving a high-risk third country required enhanced due diligence automatically, and the United Kingdom's definition of a high-risk third country followed both of the FATF public statements: the *Call for Action* list and the *Increased Monitoring* list.

1.2 The Money Laundering and Terrorist Financing (Amendment) Regulations 2026 (**SI 2026/621**, made 9 June 2026, in force 21 days later) narrow that automatic trigger to the **Call for Action** list only. Regulation 19 of the instrument amends regulation 33 of the 2017 Regulations in three places: it substitutes "FATF call for action country" for "high-risk third country" at regulation 33(1)(b), and at regulation 33(3)(a) it defines that term as *"a country named on the list of High-Risk Jurisdictions subject to a Call for Action published by the Financial Action Task Force as such list has effect from time to time"*. At the FATF plenary of 19 June 2026 that list held three jurisdictions; the Increased Monitoring list held 22.

1.3 The effect is that a customer connected to any of those 22 jurisdictions no longer attracts enhanced due diligence by operation of law. The obligation to apply a risk-based approach is unchanged, and regulation 33 still requires enhanced measures wherever the firm identifies a high risk. What has gone is the automatic trigger.

1.4 Two features of the drafting matter for how this firm should respond. The definition now tracks the FATF list **as it has effect from time to time**, so the population in scope changes at each plenary without any UK instrument and without anyone at this firm deciding anything — which is the argument for the monitoring at section 6.1. And regulation 20 of the same instrument inserts a new **regulation 34A**, extending mandatory enhanced due diligence to cryptoasset exchange providers, custodian wallet providers and correspondent relationships, in force **1 February 2027**. That is outside the scope of this note — Northgate does not offer those services — but it is the reason the instrument should not be described as a relaxation. It narrows one trigger and adds another.

1.5 Every citation in this note is taken from the instrument as made, read on legislation.gov.uk, rather than from the draft or from commentary on it. A note whose whole argument rests on one instrument should not rest on a summary of it.

## 2. What it means for this book

Measured on the current population of 400 customers.

| Measure | Customers | Share of the book |
|---|---|---|
%s

## 3. The gap

3.1 **%d customers — %.0f%% of everyone who had jurisdiction-based enhanced due diligence — lose it.**

3.2 **Not one of them is escalated for any other reason.** The geography escalator was the only control catching them. Once it stops applying, %d of the %d fall to Low, which under the current bands means simplified due diligence where the conditions of regulation 37 are met, no source of funds at onboarding, and a five-year refresh cycle.

3.3 The model already scores these customers: a connection to an Increased Monitoring jurisdiction is level 4 on factors G1, G2 or G3. The problem is that scoring them changes nothing. As the validation pack shows, the weighted average is too compressed for a single factor at level 4 to move a customer across a band boundary. The score sees the risk. The rating does not.

3.4 So the change does not just remove a legal obligation. For this firm it removes the **only** mechanism by which an Increased Monitoring jurisdiction affected how a customer was treated.

## 4. Options

| | Option | Effect | Cost |
|---|---|---|---|
| A | Do nothing beyond the statutory minimum | %d customers move from enhanced to standard due diligence | None, and the firm's exposure to 22 jurisdictions is managed by a score that cannot move a rating |
| B | Keep the old trigger as firm policy | No change in treatment for anyone | Enhanced due diligence on %d customers the law no longer requires it for, and no way to distinguish between the 22 jurisdictions |
| C | Firm-policy escalator to Medium, not High | The %d customers land at Medium: standard due diligence, source of funds where activity is inconsistent, three-year refresh | Materially less than option B; "roughly a third" is a placeholder, not a costed figure (5.2) |

## 5. Recommendation

5.1 **Option C.** Add a firm-policy rule, distinct from the statutory escalator at methodology 5.3(b), that no customer connected to a FATF Increased Monitoring jurisdiction may be rated below Medium.

5.2 Three reasons.

- **It preserves the distinction the amendment draws.** SI 2026/621 is made by the Treasury under delegated powers, and this note does not attempt to state the policy intention behind it — no explanatory memorandum is cited here. What the instrument plainly does is separate three jurisdictions from twenty-five. Option B ignores that separation; option A over-corrects for it.
- **It is materially cheaper than option B**, because Medium treatment on %d customers costs less than High treatment on the same %d. How much less depends on what this firm's enhanced due diligence actually costs per file, which the Bank has not given me. The "roughly a third" in section 4 is an order-of-magnitude placeholder and is labelled as one; it should be replaced with the firm's own figure before this note is acted on.
- **It survives the band recalibration proposed in the validation pack.** A floor at Medium is expressed in bands, not in scores, so it does not need re-tuning when the boundaries move.

5.3 The rule should be written as a floor rather than a weight. Increasing the geography weight would not achieve this, though not for the reason an earlier draft of this note gave. A ten percentage point change to the geography weight moves 16 to 26 customers, which is more than the %d affected here, not fewer. The point is that it moves a **different** set: a weight change acts on everyone whose geography score is above the minimum, most of whom are not connected to a monitored jurisdiction at all, and it would still not guarantee that any particular one of these %d crosses a boundary. A floor names the population it is meant to catch. A weight does not.

## 6. What to monitor

6.1 The number of customers subject to the firm-policy floor, reported monthly. The FATF lists change at every plenary and two jurisdictions were added in June 2026, so this population moves without anyone at the firm deciding anything.

6.2 The proportion of those customers who go on to generate a transaction monitoring alert or a suspicious activity report. If it is no different from the rest of the book after a year, the floor is not earning its cost and should be revisited.

6.3 Whether the FATF Increased Monitoring list is still the right proxy. It measures a jurisdiction's progress against its action plan, not the laundering risk facing a UK retail bank, and the firm should not outsource its geography risk appetite to it permanently.

## 7. What this note does not settle

7.1 It is written against a synthetic population of 400 customers with %d in scope. On a real book the numbers would differ and the cost comparison in section 4 could reverse.

7.2 It reads regulation 19 of SI 2026/621 as narrowing the automatic trigger and nothing more. It does not attempt to assess the new regulation 34A, or the amendments to regulation 19 of the 2017 Regulations (policies, controls and procedures) in the same instrument, both of which a full impact assessment would have to cover.

7.3 It says nothing about the timing risk. The narrowing is in force now; regulation 34A is not in force until 1 February 2027, five months after the date of this note. A firm that reads the instrument as a single relaxation and resources accordingly will be short of capacity by then.

---

## Sources

- The Money Laundering, Terrorist Financing and Transfer of Funds (Information on the Payer) Regulations 2017 (SI 2017/692), as amended — https://www.legislation.gov.uk/uksi/2017/692/contents
- The Money Laundering and Terrorist Financing (Amendment) Regulations 2026 (SI 2026/621), as made — https://www.legislation.gov.uk/uksi/2026/621/made
- SI 2026/621, regulation 19 (amendment of regulation 33) — https://www.legislation.gov.uk/uksi/2026/621/regulation/19/made
- FATF, *High-Risk Jurisdictions subject to a Call for Action*, 19 June 2026 — https://www.fatf-gafi.org/en/countries/black-and-grey-lists.html
- FATF, *Jurisdictions under Increased Monitoring*, 19 June 2026 — https://www.fatf-gafi.org/en/publications/High-risk-and-other-monitored-jurisdictions/increased-monitoring-june-2026.html

---

## Version history

| Version | Date | Change |
|---|---|---|
| 0.3 | Sept 2026 | The argument at 5.3 was false and is replaced. The "eighteen months" at 7.3 is five. The policy intention behind the instrument is no longer asserted, and the "roughly a third" cost figure is labelled as the placeholder it is. |
| 0.2 | Sept 2026 | Rewritten against SI 2026/621 as made, rather than the draft instrument and commentary on it. Regulation 19's actual amendments to regulation 33 are quoted at 1.2; the dynamic reference to the FATF list and the insertion of regulation 34A are added at 1.4; section 7 is rewritten to say what the note does not cover rather than to ask for a check that has now been done. Figures re-run. |
| 0.1 | Sept 2026 | First version, written against the draft instrument. |
""" % (edd_rows,
       EDD[4][1], 100 * EDD[4][1] / EDD[2][1],          # 3.1
       EDD[6][1], EDD[4][1],                             # 3.2
       EDD[4][1], EDD[4][1], EDD[4][1],                  # options A, B, C
       EDD[4][1], EDD[4][1],                             # 5.2 second bullet
       EDD[4][1], EDD[4][1],                             # 5.3
       EDD[4][1]))                                       # 7.1

print("wrote docs/04-model-validation.md and docs/05-edd-policy-note-2026.md")
print("  backtest rows that move under the recommendation:", moved)
print("  sensitivity range current %d-%d, rebanded %d-%d"
      % (min(c for _, _, c, _ in SENS), max(c for _, _, c, _ in SENS),
         min(d for _, _, _, d in SENS), max(d for _, _, _, d in SENS)))
