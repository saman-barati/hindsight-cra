# -*- coding: utf-8 -*-
"""Build the LinkedIn document PDF from the workbook's own figures."""
import os
from paths import REPO
import asyncio, io, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
from playwright.async_api import async_playwright
from backtest_block import CASES
from data import FACTORS

XL = REPO + "/model/customer-risk-model.xlsx"
OUT = REPO + "/hindsight-summary.pdf"
wb = load_workbook(XL, data_only=True)
V, B, W = wb["Validation"], wb["Backtest"], wb["Weights"]
sens = [(V.cell(r, 9).value, V.cell(r, 10).value) for r in range(7, 17)]
MAXA, MINB, MAXB = max(a for a, _ in sens), min(b for _, b in sens), max(b for _, b in sens)
edd = [V.cell(r, 3).value for r in range(33, 40)]

FO = [f[0] for f in FACTORS]
WEFF = {W.cell(r, 1).value: W.cell(r, 3).value for r in range(2, 22)}
NBT = sum(1 for r in range(4, 30) if B.cell(r, 1).value)
ROW = {B.cell(r, 1).value: r for r in range(4, 4 + NBT)}
def x(ref):    return B.cell(ROW[ref], 44).value
def fin(ref):  return B.cell(ROW[ref], 47).value
def sc(ref, f): return B.cell(ROW[ref], 23 + FO.index(f)).value

# the highest score any reconstruction could reach if every unsettled factor went to 5
CEIL = []
for code, name, refs in CASES:
    wp = sum(WEFF[f] for f in FO if all(sc(r, f) == 1 for r in refs))
    CEIL.append(1 * wp + 5 * (1 - wp))
MAXCEIL, MINPIN, MAXPIN = max(CEIL), None, None
pins = []
for code, name, refs in CASES:
    pins.append(sum(WEFF[f] for f in FO if all(sc(r, f) == 1 for r in refs)))
MINPIN, MAXPIN = 100 * min(pins), 100 * max(pins)
NHIGH = sum(1 for code, name, refs in CASES if any(fin(r) == "High" for r in refs))

CSS = """
@page { size: 794px 794px; margin: 0; }
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: Georgia, 'Times New Roman', serif; color: #10262C; }
.p { width: 794px; height: 794px; padding: 83px 76px; page-break-after: always;
     display: flex; flex-direction: column; position: relative; background: #F3F5F4; }
.p:last-child { page-break-after: auto; }
.p.dark { background: #10262C; color: #EAF1F0; }
.p.accent { background: #0E5563; color: #F2F7F7; }
.eyebrow { font-family: Helvetica, Arial, sans-serif; font-size: 9pt; letter-spacing: .18em;
           text-transform: uppercase; color: #0E5563; font-weight: bold; margin-bottom: 9mm; }
.dark .eyebrow, .accent .eyebrow { color: #7FC6D3; }
h1 { font-size: 40pt; line-height: 1.06; font-weight: normal; letter-spacing: -.01em; }
h2 { font-size: 26pt; line-height: 1.15; font-weight: normal; margin-bottom: 7mm; }
p  { font-size: 13.5pt; line-height: 1.55; max-width: 155mm; margin-bottom: 5mm; }
.small { font-size: 11pt; line-height: 1.5; color: #4A6268; }
.dark .small, .accent .small { color: #A9C6CB; }
.huge { font-family: Helvetica, Arial, sans-serif; font-size: 84pt; line-height: .92;
        font-weight: bold; letter-spacing: -.03em; }
.big  { font-family: Helvetica, Arial, sans-serif; font-size: 46pt; line-height: 1;
        font-weight: bold; letter-spacing: -.02em; }
.spacer { flex: 1; }
.foot { font-family: Helvetica, Arial, sans-serif; font-size: 8.5pt; letter-spacing: .1em;
        text-transform: uppercase; color: #7E9399; }
.dark .foot, .accent .foot { color: #6D949B; }
.rule { height: 2px; background: #0E5563; width: 26mm; margin: 8mm 0; }
.dark .rule, .accent .rule { background: #7FC6D3; }
table { border-collapse: collapse; font-family: Helvetica, Arial, sans-serif; font-size: 11pt; margin-top: 3mm; }
td { padding: 3.2mm 6mm 3.2mm 0; border-bottom: 1px solid #CBD8D7; }
.dark td { border-bottom: 1px solid #2A4248; }
td.n { font-weight: bold; text-align: right; padding-right: 0; }
ul { margin: 2mm 0 0 5mm; }
li { font-size: 12.5pt; line-height: 1.5; margin-bottom: 3.5mm; }
"""

PAGES = """
<div class="p dark">
  <div class="eyebrow">Financial crime compliance &middot; a self-directed project</div>
  <div class="spacer"></div>
  <h1>I built a customer<br>risk model.<br>Then I tested it on<br>the customers who<br>actually caused UK<br>banks to be fined.</h1>
  <div class="rule"></div>
  <p class="small">Saman Barati &middot; London &middot; September 2026<br>Excel and Power Query. No machine learning, on purpose.</p>
  <div class="spacer"></div>
  <div class="foot">Hindsight &middot; 1 of 9</div>
</div>

<div class="p">
  <div class="eyebrow">The idea</div>
  <h2>Most risk models are judged on whether the maths works.</h2>
  <p>That is the easy test. A customer risk assessment can be perfectly consistent and still rate the wrong people Low.</p>
  <p>So I built one properly &mdash; 20 risk factors mapped to regulation 18(2)(b) of the Money Laundering Regulations, 98 level definitions, every weight carrying a written rationale and a published source &mdash; and then rebuilt six customers named in FCA enforcement notices as they looked on the day each bank took them on.</p>
  <p><strong>Each one is scored at least twice, because a reconstruction from a published notice is twenty judgement calls and reporting one number hides nineteen of them.</strong></p>
  <p class="small">Four are read once in the customer's favour and once at the worst reading the notice does not rule out. Two are pairs of a different kind: one scored on two dates eleven days apart, and one with and without a verification the FCA found was missing.</p>
  <div class="spacer"></div>
  <div class="foot">2 of 9</div>
</div>

<div class="p accent">
  <div class="eyebrow">The result, and the problem with it</div>
  <div class="spacer"></div>
  <div class="big">Not one reaches High<br>on the arithmetic.</div>
  <p style="font-size:15pt; margin-top:7mm; max-width:150mm;">In any reading. Three of the six are rated High, and every one of them through the same mandatory escalator rather than through a score.</p>
  <div class="rule"></div>
  <div class="huge">%.2f</div>
  <p style="font-size:15pt; margin-top:5mm; max-width:150mm;">is the highest score any of them <em>could</em> have reached, with every open factor pushed to 5. High begins at 3.50.</p>
  <p class="small" style="margin-top:6mm;">The facts each notice settles &mdash; a UK company with UK owners, or an ordinary UK personal customer &mdash; pin %.0f&ndash;%.0f%% of the model's weight at level 1 before any judgement is made. The headline finding was fixed by the inputs, not found by the test. What the back-test can still decide is whether an escalator fires.</p>
  <div class="spacer"></div>
  <div class="foot">3 of 9</div>
</div>

<div class="p">
  <div class="eyebrow">Finding one</div>
  <h2>One word decided the most serious case.</h2>
  <p>Fowler Oldfield was a Bradford jewellery business. Around &pound;365m was deposited with NatWest, roughly &pound;264m of it in cash. It produced the FCA's first criminal prosecution under the money laundering regulations.</p>
  <table>
    <tr><td>Read as a <strong>cash&#8209;intensive trade</strong></td><td class="n">%.2f &rarr; Low</td></tr>
    <tr><td>Read as a <strong>dealer in high&#8209;value goods</strong></td><td class="n">%.2f &rarr; High</td></tr>
  </table>
  <p style="margin-top:6mm;">Same file, same day. The scores differ by 0.075 and neither crosses a band. The rating differs completely, because one reading fires an escalator and the other does not.</p>
  <p class="small">That is a defect in my level definitions, not in the analyst reading them. So I wrote a fix.</p>
  <div class="spacer"></div>
  <div class="foot">4 of 9</div>
</div>

<div class="p dark">
  <div class="eyebrow">Finding two &middot; the one I did not want</div>
  <h2>My fix would have broken the only case the model got right.</h2>
  <p>Barclays opened an account for a gold refining and trading company in January 2015. The application form said &pound;500,000 turnover, manually amended to &pound;3 million, and that it would not trade outside the EU. Eleven days later the file recorded gold from Ghana and Burkina Faso.</p>
  <p><strong>My model rates it High on day one</strong> &mdash; %.2f, escalated &mdash; because "gold refining and trading" is a dealer in high-value goods under the same ambiguous definition that page four calls a defect.</p>
  <p>Tie that level to <em>registration</em>, as I had recommended, and this customer drops to Medium. A gold trader declaring no cash would not have been registered.</p>
  <p class="small">I wrote a fix, tested it on the case that motivated it, and never asked what else it touched. The recommendation is still on the list, now with the argument against it attached.</p>
  <div class="spacer"></div>
  <div class="foot">5 of 9</div>
</div>

<div class="p">
  <div class="eyebrow">Finding three</div>
  <h2>A customer whose address was a London landmark scores %.2f.</h2>
  <p>Monzo onboarded customers who gave well&#8209;known London landmarks as their home address. My library has a factor written for exactly that &mdash; from that notice, before any of it was scored.</p>
  <p>I then gave the factor <strong>1.5%% of the model</strong>, and wrote a paragraph defending the weight.</p>
  <p>Scored through, the customer lands comfortably in Low. The factor that saw the problem was too light to change the answer.</p>
  <p class="small">Both the reasoning and the result stay in the repository. Deleting the argument I got wrong would have removed the only evidence that the test worked.</p>
  <div class="spacer"></div>
  <div class="foot">6 of 9</div>
</div>

<div class="p">
  <div class="eyebrow">Finding four</div>
  <h2>The weights were never the thing that mattered.</h2>
  <p>I spent a whole document arguing about whether geography should carry 25%% or 15%%.</p>
  <table>
    <tr><td>Move any category weight by 10 percentage points</td><td class="n">up to %d of 400<br>customers change band</td></tr>
    <tr><td>The same change, against band boundaries<br>placed where the population actually sits</td><td class="n">%d to %d<br>customers change band</td></tr>
  </table>
  <p style="margin-top:7mm;">The boundaries decide the output. The weights barely touch it.</p>
  <p class="small">Which reverses the order of the work: fix the bands first, argue about the weights second. Those boundaries are fitted to a population I invented, which is circular, and the recommendation says so where it is made rather than in a footnote.</p>
  <div class="spacer"></div>
  <div class="foot">7 of 9</div>
</div>

<div class="p">
  <div class="eyebrow">Finding five &middot; a live question</div>
  <h2>The 2026 EDD change removes the only control catching a whole population.</h2>
  <p>SI 2026/621, made 9 June 2026, narrows mandatory jurisdiction&#8209;based enhanced due diligence to the three FATF <em>Call for Action</em> countries, rather than those plus the 22 under <em>Increased Monitoring</em>.</p>
  <table>
    <tr><td>Had jurisdiction&#8209;based EDD before 2026</td><td class="n">%d</td></tr>
    <tr><td>Have it under the 2026 rule</td><td class="n">%d</td></tr>
    <tr><td>Lose it</td><td class="n">%d</td></tr>
    <tr><td><strong>Of those, escalated for any other reason</strong></td><td class="n">%d</td></tr>
  </table>
  <p style="margin-top:6mm;">The model still scores them. It just cannot move them across a band. The score sees the risk; the rating does not.</p>
  <div class="spacer"></div>
  <div class="foot">8 of 9</div>
</div>

<div class="p accent">
  <div class="eyebrow">Finding six &middot; the one I would keep</div>
  <h2>The model had the control. The file did not have the input.</h2>
  <p>Santander opened a business account on a form describing the trade as a &ldquo;Translation service&rdquo;, expecting &pound;5,000 a month. Within six months it was receiving millions.</p>
  <table>
    <tr><td>Scored from the application form</td><td class="n">%.2f &rarr; Low</td></tr>
    <tr><td>With the business verified as an <strong>MSB</strong></td><td class="n">%.2f &rarr; High</td></tr>
  </table>
  <p style="margin-top:6mm;">The Final Notice records that the customer operated a money service business and that this was not identified at onboarding. Verified, the escalator fires on day one. The control was in the model the whole time.</p>
  <p class="small">Five documents of this project argue about weights, boundaries and aggregation rules. None of that argument survives contact with a file where the trade was never verified. A risk model is only ever as good as the due diligence feeding it, and nothing in mine measures that.</p>
  <div class="spacer"></div>
  <p class="small"><strong>Synthetic data, a fictional bank, and a self&#8209;directed learning project alongside a BA in Accounting and Finance. Not professional advice, and not independently validated by a person.</strong> Full method, workbooks and case files on GitHub &mdash; link in the comments.</p>
  <div class="foot">9 of 9 &middot; Saman Barati &middot; London</div>
</div>
""" % (MAXCEIL, MINPIN, MAXPIN,
         x("FO-lo"), x("FO-mid"),
         x("STU-a"),
         x("MON-lo"),
         MAXA, MINB, MAXB,
         edd[2], edd[3], edd[4], edd[5],
         x("SAN-a"), x("SAN-b"))

HTML = "<!doctype html><html><head><meta charset=\"utf-8\"><style>%s</style></head><body>%s</body></html>" % (CSS, PAGES)


async def main():
    io.open("/tmp/li.html", "w", encoding="utf-8").write(HTML)
    async with async_playwright() as pw:
        br = await pw.chromium.launch(executable_path="/opt/pw-browsers/chromium")
        pg = await br.new_page()
        await pg.goto("file:///tmp/li.html")
        await pg.pdf(path=OUT, width="794px", height="794px", print_background=True,
                     margin={"top": "0", "bottom": "0", "left": "0", "right": "0"})
        await br.close()
    print("wrote", OUT)

asyncio.run(main())
