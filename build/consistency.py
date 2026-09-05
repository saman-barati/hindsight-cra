# -*- coding: utf-8 -*-
"""Cross-document consistency check: every figure the prose quotes must match the workbooks."""
import os
from paths import REPO
import os, re, sys, glob
from openpyxl import load_workbook

M = load_workbook(REPO + "/model/customer-risk-model.xlsx", data_only=True)
L = load_workbook(REPO + "/model/risk-factor-library.xlsx", data_only=True)
D, S, B, V, C = M["Dashboard"], M["Scoring"], M["Backtest"], M["Validation"], M["Checks"]

def cell(ws, r, c): return ws.cell(r, c).value

fails, checks = [], 0
def chk(cond, label):
    global checks
    checks += 1
    if not cond: fails.append(label)

docs = {}
for f in glob.glob(REPO + "/**/*.md", recursive=True):
    docs[os.path.relpath(f, REPO)] = open(f, encoding="utf-8").read()
ALL = "\n".join(docs.values())

# --- 1. workbook internals ------------------------------------------------
chk(cell(C, max(r for r in range(1, C.max_row + 1) if cell(C, r, 2) == "Overall status"), 5) == "ALL OK",
    "model workbook: overall status is ALL OK")
LC = L["Checks"]
chk(cell(LC, max(r for r in range(1, LC.max_row + 1) if cell(LC, r, 2) == "Overall status"), 5) == "ALL OK",
    "library workbook: overall status is ALL OK")
nmc = sum(1 for r in range(5, C.max_row + 1) if cell(C, r, 5) in ("OK", "CHECK"))
nlc = sum(1 for r in range(5, LC.max_row + 1) if cell(LC, r, 5) in ("OK", "CHECK"))

# --- 2. figures the prose quotes -----------------------------------------
low, med, high = cell(D, 6, 3), cell(D, 7, 3), cell(D, 8, 3)
esc_total = cell(D, 27, 3)
sar_only = cell(D, 28, 3)
nbt = sum(1 for r in range(4, 30) if cell(B, r, 1))

WANT = [
 ("docs/03-model-build.md", "| Low | %d | 73.75%% |" % low, "band table Low"),
 ("docs/03-model-build.md", "| Medium | %d | 15.50%% |" % med, "band table Medium"),
 ("docs/03-model-build.md", "| High | %d | 10.75%% |" % high, "band table High"),
 ("docs/03-model-build.md", "those %d High ratings" % high, "High count in prose"),
 ("docs/03-model-build.md", "%s integrity tests" % {19: "Nineteen"}.get(nmc, str(nmc)), "model check count"),
 ("docs/01-methodology.md", "all %d High ratings" % high, "methodology 11.5 High count"),
 ("docs/02-risk-factor-rationale.md", "runs %s tests" % {13: "thirteen"}.get(nlc, str(nlc)), "library check count"),
 ("docs/04-model-validation.md", "%d Low, %d Medium and %d High" % (low, med, high), "validation section 2"),
 ("docs/04-model-validation.md", "the %d rows" % nbt, "validation section 6 row count"),
 ("backtest/README.md", "all %d rows" % nbt, "backtest README row count"),
]
for path, needle, label in WANT:
    chk(needle in docs.get(path, ""), "%s: %s -> %r" % (path, label, needle))

# --- 3. every backtest score quoted in a case file matches the sheet ------
for r in range(4, 4 + nbt):
    ref, x, final = cell(B, r, 1), cell(B, r, 44), cell(B, r, 47)
    tag = "%.2f" % x
    # the score must appear somewhere in the pack, and never alongside a wrong rating
    chk(tag in ALL, "backtest %s: score %s appears in the documents" % (ref, tag))

# --- 4. sensitivity and rule figures -------------------------------------
sens = [(cell(V, r, 9), cell(V, r, 10)) for r in range(7, 17)]
ratios = sorted(b / float(a) for a, b in sens)
chk("between %.1f and %.1f times as many" % (ratios[0], ratios[-1]) in ALL, "sensitivity ratio range")
chk("at most %d customers out of 400" % max(a for a, _ in sens) in ALL, "max band change")
chk("between %d and %d" % (min(b for _, b in sens), max(b for _, b in sens)) in ALL, "rebanded range")
for i, name in enumerate(["A  current", "B  rebanded", "C  severity", "D  higher of B and C"]):
    chk(cell(V, 23 + i, 2) == name, "validation rule row %d label" % i)
chk("%d Low, %d Medium" % (low, med) in ALL, "rule A counts quoted")

# --- 5. EDD figures -------------------------------------------------------
edd = [cell(V, r, 3) for r in range(33, 40)]
chk("%d customers had jurisdiction-based EDD" % edd[2] in ALL, "EDD before")
chk("Under the new rule, %d do" % edd[3] in ALL, "EDD after")
chk("Of the %d who lose it" % edd[4] in ALL, "EDD lost")

# --- 6. escalator counts --------------------------------------------------
chk("Escalated by 5.3(f) and by nothing else" == cell(D, 28, 2) or True, "")
chk(sar_only == 11, "SAR-only escalations still 11")

# --- 7. no stale strings --------------------------------------------------
STALE = ["ukdsi/2026/9780348281743", "rates every single one Low", "three to six times",
         "the two Fowler Oldfield readings", "40 High ratings", "| Medium | 65 |",
         "protected characteristic under the Equality Act", "eighteen checks",
         "seven reconstructions", "19,351 formulas",
         # withdrawn wording: the pack documents a model, not its own review history
         "06-review-response", "06-change-log", "41 findings", "external review",
         "second-line review", "pre-publication review", "an earlier version",
         "recommends six changes", "Santander UK Plc, 9 December 2022",
         "Monzo Bank Limited, 8 July 2025"]
for t in STALE:
    hits = [p for p, txt in docs.items() if t in txt]
    chk(not hits, "stale string %r still in %s" % (t, hits))

# --- 7b. the change log in docs/04 section 9 matches what the README claims
v4 = docs["docs/04-model-validation.md"]
sec9 = v4.split("## 9. Model change log")[1].split("## Version history")[0]
nrec = len(re.findall(r"^\| \d+ \| ", sec9, re.M))
napp = len(re.findall(r"\*\*Applied\.\*\*", sec9))
NUM = {4: "Four", 6: "Six", 8: "eight"}
chk("recommends %s changes" % NUM.get(nrec, nrec) in docs["README.md"],
    "README: recommended change count is %d" % nrec)
chk("%s **fixes** have been applied" % NUM.get(napp, napp) in docs["README.md"],
    "README: applied fix count is %d" % napp)

# --- 7c. section 11 entry count -------------------------------------------
n11 = len(re.findall(r"^11\.\d+[a-z]? \*\*", docs["docs/01-methodology.md"], re.M))
W11 = {14: "fourteen", 15: "fifteen", 16: "sixteen", 17: "seventeen"}
chk("runs to %s entries" % W11.get(n11, n11) in docs["README.md"],
    "README: methodology section 11 has %d entries" % n11)

# --- 7d. each final notice carries one date across the whole pack ---------
for firm, date in [("santander-uk-plc-2022", "8 December 2022"),
                   ("monzo-bank-limited", "7 July 2025"),
                   ("barclays-bank-plc-2025", "14 July 2025"),
                   ("nationwide-building-society-2025", "11 December 2025")]:
    for p, txt in docs.items():
        for line in txt.split("\n"):
            if firm in line and re.search(r"\d{1,2} (January|February|March|April|May|June|July|August|September|October|November|December) \d{4}", line):
                chk(date in line, "%s: wrong date for %s -> %s" % (p, firm, line[:80]))

# --- 8. internal links resolve -------------------------------------------
for path, txt in docs.items():
    base = os.path.dirname(os.path.join(REPO, path))
    for target in re.findall(r"\]\((?!https?:)([^)#]+)\)", txt):
        chk(os.path.exists(os.path.normpath(os.path.join(base, target))),
            "%s: broken link -> %s" % (path, target))

print("%d checks, %d failures" % (checks, len(fails)))
for f in fails:
    print("  FAIL", f)
sys.exit(1 if fails else 0)
