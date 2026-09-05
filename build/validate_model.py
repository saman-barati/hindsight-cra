# -*- coding: utf-8 -*-
"""Independent validation of Step 3.

Recomputes all 400 ratings in Python straight from the CSV and the Step 2
library, then requires an exact match against what the workbook itself
calculated. If the two disagree anywhere, the workbook is wrong.
"""
import os
from paths import REPO
import sys, csv, statistics
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data import CATEGORIES, FACTORS
from labels import LABELS
from openpyxl import load_workbook

CSV = REPO + "/data/synthetic-customers.csv"
XL = REPO + "/model/customer-risk-model.xlsx"
LIB = REPO + "/model/risk-factor-library.xlsx"

FACTOR_ORDER = [f[0] for f in FACTORS]
CAT_OF = {f[0]: f[1] for f in FACTORS}
WIC = {f[0]: f[3] for f in FACTORS}
CATW = {c[0]: c[2] for c in CATEGORIES}
SCORE_OF = {f: {lab: s for s, lab in LABELS[f].items()} for f in FACTOR_ORDER}

fails = []
def chk(cond, msg):
    print(("  OK   " if cond else "  FAIL ") + msg)
    if not cond:
        fails.append(msg)

def band(x):
    return "Low" if x <= 2 else ("Medium" if x <= 3.49 else "High")

rows = list(csv.DictReader(open(CSV, encoding="utf-8")))

# ---- independent recomputation -------------------------------------------
expected = []
for row in rows:
    sc = {f: SCORE_OF[f][row[f]] for f in FACTOR_ORDER}
    cats = {}
    for f in FACTOR_ORDER:
        cats[CAT_OF[f]] = cats.get(CAT_OF[f], 0.0) + WIC[f] * sc[f]
    overall = sum(CATW[c] * cats[c] for c in cats)
    esc = (sc["C4"] >= 3 or sc["C3"] == 5 or sc["C2"] == 5 or sc["C1"] == 5
           or sc["C5"] == 5
           or sc["G1"] == 5 or sc["G2"] == 5 or sc["G3"] == 5
           or row["sar_last_12m"] == "Yes")
    b = band(round(overall, 2))
    expected.append({"id": row["customer_id"], "segment": row["segment"], "scores": sc,
                     "cats": cats, "overall": overall, "band": b,
                     "esc": "Yes" if esc else "No",
                     "final": "High" if esc else b})

# ---- what the workbook says ----------------------------------------------
wb = load_workbook(XL, data_only=True)
ws = wb["Scoring"]
n = len(expected)

print("SCORING ENGINE vs INDEPENDENT RECOMPUTATION (%d customers)" % n)
mismatch_scores = mismatch_overall = mismatch_band = mismatch_esc = mismatch_final = mismatch_cat = 0
for i, e in enumerate(expected):
    r = 2 + i
    if ws.cell(r, 1).value != e["id"]:
        mismatch_scores += 1
    for j, f in enumerate(FACTOR_ORDER):
        if ws.cell(r, 3 + j).value != e["scores"][f]:
            mismatch_scores += 1
    if round(ws.cell(r, 23).value, 9) != round(e["overall"], 9):
        mismatch_overall += 1
    if ws.cell(r, 25).value != e["band"]:
        mismatch_band += 1
    if ws.cell(r, 26).value != e["esc"]:
        mismatch_esc += 1
    if ws.cell(r, 27).value != e["final"]:
        mismatch_final += 1
    for k, code in enumerate(["C", "G", "P", "D", "A"]):
        if round(ws.cell(r, 28 + k).value, 9) != round(e["cats"][code], 9):
            mismatch_cat += 1

chk(mismatch_scores == 0, "every factor score matches (%d mismatches)" % mismatch_scores)
chk(mismatch_overall == 0, "every overall score matches (%d mismatches)" % mismatch_overall)
chk(mismatch_cat == 0, "every category score matches (%d mismatches)" % mismatch_cat)
chk(mismatch_band == 0, "every arithmetic band matches (%d mismatches)" % mismatch_band)
chk(mismatch_esc == 0, "every escalator flag matches (%d mismatches)" % mismatch_esc)
chk(mismatch_final == 0, "every final rating matches (%d mismatches)" % mismatch_final)

# ---- weights match the Step 2 library ------------------------------------
lib = load_workbook(LIB, data_only=True)["Factors"]
lib_eff = {lib.cell(r, 1).value: round(lib.cell(r, 5).value, 10) for r in range(2, 22)}
wsw = wb["Weights"]
model_eff = {wsw.cell(r, 1).value: round(wsw.cell(r, 3).value, 10) for r in range(2, 22)}
chk(lib_eff == model_eff, "effective weights identical to the Step 2 library")

# ---- workbook checks sheet ------------------------------------------------
wsc = wb["Checks"]
print("\nWORKBOOK CHECKS SHEET")
for r in range(5, 17):
    if wsc.cell(r, 1).value is None:
        continue
    chk(wsc.cell(r, 5).value == "OK",
        "check %s: %s (result %s, expected %s)" % (wsc.cell(r, 1).value, str(wsc.cell(r, 2).value)[:52],
                                                   wsc.cell(r, 3).value, wsc.cell(r, 4).value))
st_row = max(r for r in range(1, wsc.max_row + 1) if wsc.cell(r, 2).value == "Overall status")
n_checks = sum(1 for r in range(5, st_row) if wsc.cell(r, 5).value in ("OK", "CHECK"))
n_fail = sum(1 for r in range(5, st_row) if wsc.cell(r, 5).value == "CHECK")
chk(wsc.cell(st_row, 5).value == "ALL OK", "overall status reads ALL OK (%d checks, %d failing)" % (n_checks, n_fail))

# ---- headline results -----------------------------------------------------
print("\nRESULTS (from the independent recomputation)")
finals = [e["final"] for e in expected]
bands = [e["band"] for e in expected]
for b in ["Low", "Medium", "High"]:
    print("  final %-7s %3d  (%4.1f%%)   arithmetic-only %3d (%4.1f%%)"
          % (b, finals.count(b), 100 * finals.count(b) / n, bands.count(b), 100 * bands.count(b) / n))
lifted = sum(1 for e in expected if e["band"] != "High" and e["final"] == "High")
print("  lifted to High by an escalator: %d" % lifted)
print("  escalated customers:            %d" % sum(1 for e in expected if e["esc"] == "Yes"))

print("\nCATEGORY DISCRIMINATION")
print("  %-20s %6s %6s %6s %8s %10s" % ("category", "mean", "min", "max", "stdev", "at min"))
for code, name, w in CATEGORIES:
    vals = [e["cats"][code] for e in expected]
    atmin = sum(1 for v in vals if abs(v - min(vals)) < 1e-9) / n
    print("  %-20s %6.2f %6.2f %6.2f %8.3f %9.1f%%   (weight %.0f%%)"
          % (name, statistics.mean(vals), min(vals), max(vals),
             statistics.pstdev(vals), 100 * atmin, 100 * w))

ov = [e["overall"] for e in expected]
print("\nOVERALL SCORE  mean %.3f  min %.2f  max %.2f  stdev %.3f"
      % (statistics.mean(ov), min(ov), max(ov), statistics.pstdev(ov)))
near = sum(1 for e in expected if 1.91 <= round(e["overall"], 2) <= 2.10)
print("  within 0.10 of the Low/Medium boundary: %d (%.1f%%)" % (near, 100 * near / n))

print("\nBY SEGMENT")
for seg in ["Personal", "Sole trader", "Limited company"]:
    sub = [e for e in expected if e["segment"] == seg]
    print("  %-16s n=%3d  Low %3d  Medium %3d  High %3d"
          % (seg, len(sub), sum(1 for e in sub if e["final"] == "Low"),
             sum(1 for e in sub if e["final"] == "Medium"),
             sum(1 for e in sub if e["final"] == "High")))

# cash-intensive sole traders, the 11.4 question
ci = [e for e in expected if e["scores"]["C3"] == 4 and e["scores"]["A2"] >= 4 and e["esc"] == "No"]
print("\nCASH-INTENSIVE, HIGH-CASH, NOT ESCALATED: %d customers" % len(ci))
if ci:
    print("  their ratings: Low %d, Medium %d, High %d"
          % (sum(1 for e in ci if e["final"] == "Low"),
             sum(1 for e in ci if e["final"] == "Medium"),
             sum(1 for e in ci if e["final"] == "High")))
    print("  highest score among them: %.2f" % max(e["overall"] for e in ci))

print("\nRESULT:", "ALL CHECKS PASSED" if not fails else "%d FAILURES" % len(fails))
