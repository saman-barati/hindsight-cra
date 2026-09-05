# -*- coding: utf-8 -*-
"""Assemble review-pack.md: every document in the project plus an appendix of
figures read straight out of the workbooks. Regenerate this before any re-review."""
import os
from paths import REPO
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
from data import FACTORS
from backtest_block import CASES

OUT = REPO + "/../review-pack.md"
M = load_workbook(REPO + "/model/customer-risk-model.xlsx", data_only=True)
L = load_workbook(REPO + "/model/risk-factor-library.xlsx", data_only=True)
D, S, B, V, C, W = M["Dashboard"], M["Scoring"], M["Backtest"], M["Validation"], M["Checks"], M["Weights"]
LF, LL, LE, LC = L["Factors"], L["Levels"], L["Escalators"], L["Checks"]

FO = [f[0] for f in FACTORS]
NAME = {f[0]: f[2] for f in FACTORS}
NBT = sum(1 for r in range(4, 40) if B.cell(r, 1).value)
out = []
w = out.append


def sheet_table(ws, r0, ncol, stop_col=1):
    rows = []
    r = r0
    while ws.cell(r, stop_col).value not in (None, ""):
        rows.append([ws.cell(r, c).value for c in range(1, ncol + 1)])
        r += 1
    return rows


def fmt(v):
    if v is None: return ""
    if isinstance(v, float):
        return ("%.4f" % v).rstrip("0").rstrip(".") if abs(v) < 1 else "%.4g" % v
    return str(v)


def table(head, rows):
    w("| " + " | ".join(head) + " |")
    w("|" + "---|" * len(head))
    for r in rows:
        w("| " + " | ".join(fmt(x) for x in r) + " |")
    w("")


w("# Hindsight — full review pack")
w("")
w("This single file contains every written document in the project, in the order it was produced, "
  "plus an appendix of the numbers the documents quote, taken directly from the workbooks.")
w("")
w("The two Excel workbooks and the customer CSV are separate files. If you cannot open them, "
  "Appendix A below carries the figures they contain.")
w("")
w("---")
w("")
w("# Appendix A — the numbers, straight from the workbooks")
w("")
w("Every figure below was read out of the spreadsheets, not retyped from the prose. If a document "
  "states a number that disagrees with this appendix, the document is wrong.")
w("")

w("## A1. The 20 risk factors and their weights")
w("")
table(["Code", "Category", "Factor", "Weight in category", "Effective weight"],
      [[W.cell(r, 1).value, W.cell(r, 2).value, NAME[W.cell(r, 1).value],
        "%.2f%%" % (100 * W.cell(r, 4).value), "%.2f%%" % (100 * W.cell(r, 3).value)]
       for r in range(2, 22)])

w("## A2. All 98 level definitions")
w("")
rows = []
r = 2
while LL.cell(r, 1).value:
    rows.append([LL.cell(r, c).value for c in range(1, 6)])
    r += 1
table(["Factor", "Level", "Definition", "Escalator", "Source"] if len(rows[0]) == 5 else
      ["Factor", "Level", "Definition", "Escalator"], rows)

w("## A3. The 400-customer population, scored")
w("")
w("- Customers: %d" % D.cell(9, 3).value)
w("- Overall score: min %.2f, mean %.3f, max %.2f" % (D.cell(51, 3).value, D.cell(50, 3).value, D.cell(52, 3).value))
w("- Final ratings: Low %d, Medium %d, High %d" % (D.cell(6, 3).value, D.cell(7, 3).value, D.cell(8, 3).value))
w("- Band from the arithmetic alone: Low %d, Medium %d, High %d"
  % (D.cell(13, 3).value, D.cell(14, 3).value, D.cell(15, 3).value))
w("- Customers within 0.10 of the Low/Medium boundary: %d" % D.cell(54, 3).value)
w("")
w("Why customers were escalated:")
w("")
table(["Escalator", "Customers"], [[D.cell(r, 2).value, D.cell(r, 3).value] for r in range(21, 29)])
w("Category statistics across the population:")
w("")
table(["Category", "Mean", "Min", "Max", "Std dev", "Share at the category minimum"],
      [[D.cell(r, 2).value, "%.2f" % D.cell(r, 3).value, "%.2f" % D.cell(r, 4).value,
        "%.2f" % D.cell(r, 5).value, "%.3f" % D.cell(r, 6).value, "%.2f%%" % (100 * D.cell(r, 7).value)]
       for r in range(40, 45)])
w("Rating by segment:")
w("")
table(["Segment", "Low", "Medium", "High", "Total"],
      [[D.cell(r, 2).value, D.cell(r, 3).value, D.cell(r, 4).value, D.cell(r, 5).value, D.cell(r, 6).value]
       for r in range(33, 36)])

w("## A4. Weight sensitivity")
w("")
table(["Scenario", "Effect", "C", "G", "P", "D", "A", "Band changes, current bands", "Band changes, rebanded"],
      [[V.cell(r, c).value for c in range(2, 11)] for r in range(7, 17)])

w("## A5. Aggregation rules compared, on the same 400 customers")
w("")
table(["Rule", "Definition", "Low", "Medium", "High", "High share"],
      [[V.cell(r, 2).value, V.cell(r, 3).value, V.cell(r, 4).value, V.cell(r, 5).value,
        V.cell(r, 6).value, "%.1f%%" % (100 * V.cell(r, 7).value)] for r in range(23, 27)])

w("## A6. The 2026 EDD test")
w("")
table(["Measure", "Customers", "Share of the book"],
      [[V.cell(r, 2).value, V.cell(r, 3).value, "%.2f%%" % (100 * V.cell(r, 4).value)] for r in range(33, 40)])

w("## A7. The %d back-test rows, with every input" % NBT)
w("")
table(["Case"] + FO,
      [[B.cell(r, 1).value] + [B.cell(r, 3 + i).value for i in range(20)] for r in range(4, 4 + NBT)])
w("Scores for each:")
w("")
table(["Case", "Reconstruction", "C", "G", "P", "D", "A", "Weighted", "Rounded", "Band",
       "Escalator", "Final", "Rule B", "Recommended"],
      [[B.cell(r, 1).value, B.cell(r, 2).value] + ["%.2f" % B.cell(r, c).value for c in range(48, 53)]
       + ["%.4f" % B.cell(r, 43).value, "%.2f" % B.cell(r, 44).value]
       + [B.cell(r, c).value for c in (45, 46, 47, 53, 54)] for r in range(4, 4 + NBT)])
w("Which rows belong to which case:")
w("")
table(["Case", "Rows"], [[n, ", ".join(refs)] for c, n, refs in CASES])

w("## A8. The workbooks' own integrity checks")
w("")
w("Model workbook:")
w("")
table(["#", "Test", "Result", "Expected", "Status"],
      [[C.cell(r, c).value for c in range(1, 6)] for r in range(5, C.max_row + 1)
       if C.cell(r, 5).value in ("OK", "CHECK")])
w("Library workbook:")
w("")
table(["#", "Test", "Result", "Expected", "Status"],
      [[LC.cell(r, c).value for c in range(1, 6)] for r in range(5, LC.max_row + 1)
       if LC.cell(r, 5).value in ("OK", "CHECK")])

w("## A9. The mandatory escalators and prohibitions, as the library states them")
w("")
rows, r = [], 2
while LE.cell(r, 1).value:
    rows.append([LE.cell(r, c).value for c in range(1, 4)])
    r += 1
table(["Ref", "Effect", "Condition"], rows)

FILES = [
 ("README.md", "The repository front page. What the project claims."),
 ("docs/01-methodology.md", "Step 1. The firm, the perimeter, the factors, the scoring rules, the governance."),
 ("docs/02-risk-factor-rationale.md", "Step 2. Why each factor and weight was chosen."),
 ("docs/03-model-build.md", "Step 3. Building the workbook and the first run on 400 customers."),
 ("data/generation-notes.md", "How the synthetic population was generated, and its limitations."),
 ("backtest/README.md", "Step 4. The back-test against six FCA enforcement cases."),
 ("backtest/cases/natwest-fowler-oldfield.md", "Back-test case file."),
 ("backtest/cases/santander-translations-company.md", "Back-test case file."),
 ("backtest/cases/barclays-stunt-and-co.md", "Back-test case file."),
 ("backtest/cases/barclays-wealthtek.md", "Back-test case file."),
 ("backtest/cases/monzo-landmark-address.md", "Back-test case file."),
 ("backtest/cases/nationwide-business-on-personal.md", "Back-test case file."),
 ("docs/04-model-validation.md", "Step 5. Distribution, sensitivity, aggregation rules, recommendations."),
 ("docs/05-edd-policy-note-2026.md", "Step 5. The policy note on SI 2026/621."),
]
for path, blurb in FILES:
    w("")
    w("=" * 78)
    w("")
    w("# FILE: `%s`" % path)
    w("")
    w("*%s*" % blurb)
    w("")
    w("-" * 78)
    w("")
    w(open(REPO + "/" + path, encoding="utf-8").read().rstrip())

open(OUT, "w", encoding="utf-8").write("\n".join(out) + "\n")
print("wrote %s (%d lines, %d words)" % (OUT, len(out), len(" ".join(out).split())))
