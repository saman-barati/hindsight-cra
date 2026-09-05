# -*- coding: utf-8 -*-
import os
from paths import REPO
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data import CATEGORIES, FACTORS, LEVELS, ESCALATORS, EXAMPLE
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

ARIAL   = "Arial"
HDR_FILL = PatternFill("solid", fgColor="14454E")
SUB_FILL = PatternFill("solid", fgColor="E4EAEA")
IN_FILL  = PatternFill("solid", fgColor="FFF2CC")
OK_FILL  = PatternFill("solid", fgColor="E2EFDA")
H  = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
B  = Font(name=ARIAL, size=10, bold=True)
N  = Font(name=ARIAL, size=10)
IN = Font(name=ARIAL, size=10, color="0000FF")
MU = Font(name=ARIAL, size=9, color="5C6E72")
TITLE = Font(name=ARIAL, size=14, bold=True, color="14454E")
TOP  = Alignment(vertical="top")
WRAP = Alignment(vertical="top", wrap_text=True)
CTR  = Alignment(vertical="top", horizontal="center")
THIN = Side(style="thin", color="C9D4D4")
BOX  = Border(bottom=THIN)

NF = len(FACTORS)          # 20
NL = len(LEVELS)           # 98
FR = (2, 1 + NF)           # factor rows 2..21
LR = (2, 1 + NL)           # level rows  2..99
CR = (2, 1 + len(CATEGORIES))

wb = Workbook()

def header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill, c.alignment = H, HDR_FILL, Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28

def add_table(ws, name, first_row, last_row, last_col):
    ref = "A%d:%s%d" % (first_row, get_column_letter(last_col), last_row)
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight8", showRowStripes=True)
    ws.add_table(t)

# ----------------------------------------------------------------- Overview
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 104

rows = [
 (TITLE, "Risk factor library"),
 (B,     "Northgate Bank UK Limited (fictional)  |  HND-CRA-002  |  version 0.3 draft  |  September 2026"),
 (N,     ""),
 (B,     "What this workbook is"),
 (N,     "Step 2 of the Hindsight project. It turns the framework set out in docs/01-methodology.md into a working library: "
         "20 risk factors across the five categories required by regulation 18(2)(b) of the Money Laundering Regulations 2017, "
         "each with defined levels, a score from 1 to 5, a weight, a written rationale and a source."),
 (N,     ""),
 (B,     "How to read it"),
 (N,     "Categories  -  the five headings and their weights. These sum to 100%."),
 (N,     "Factors     -  the 20 factors, their weight inside their category, and their effective weight across the whole model."),
 (N,     "Levels      -  what each score from 1 to 5 means for each factor. An analyst selects the level that matches the evidence on file."),
 (N,     "Escalators  -  the mandatory High and prohibited lists from sections 5.3 and 5.4 of the methodology."),
 (N,     "Example     -  a worked customer, scored live from the library."),
 (N,     "Checks      -  arithmetic and integrity tests. Every one must read OK before the workbook is used."),
 (N,     ""),
 (B,     "Which cells to edit"),
 (N,     "Cells in blue text on a pale yellow fill are inputs: the category weights on Categories, the factor weights on Factors, "
         "and the selected scores on Example. Everything else is a formula and should not be overwritten. Changing any weight "
         "recalculates the whole workbook, including the Checks sheet."),
 (N,     ""),
 (B,     "Scoring rules"),
 (N,     "1.  Each factor is scored on the level definitions in the Levels sheet. Analysts select a level; they do not choose a number."),
 (N,     "2.  Where a customer's declared figure straddles two levels, the higher level applies."),
 (N,     "3.  Category score = sum of (factor weight x factor score) within the category."),
 (N,     "4.  Overall score = sum of (category weight x category score), rounded to two decimal places."),
 (N,     "5.  Bands: Low 1.00 to 2.00, Medium 2.01 to 3.49, High 3.50 to 5.00."),
 (N,     "6.  A mandatory escalator in section 5.3 rates the customer High whatever the arithmetic produces."),
 (N,     ""),
 (B,     "Assumptions that are not drawn from a published source"),
 (N,     "The category weights and the factor weights are judgement. So are the monetary thresholds in P2 and A1 and the percentage "
         "bands in A2 and A3. Each is marked in the Source column of the Factors sheet. Step 5 tests how far the output moves when "
         "each weight is changed; until that is done the weights should be treated as provisional."),
 (N,     ""),
 (B,     "Sources"),
 (N,     "The Money Laundering, Terrorist Financing and Transfer of Funds (Information on the Payer) Regulations 2017 (SI 2017/692), as amended"),
 (N,     "JMLSG, Guidance for the UK financial sector, Part I, chapters 4 and 5"),
 (N,     "FCA Financial Crime Guide (FCG); FCA FG25/3 on politically exposed persons (July 2025), which replaces FG17/6"),
 (N,     "HM Treasury and Home Office, National Risk Assessment of Money Laundering and Terrorist Financing 2025 (July 2025)"),
 (N,     "FATF, High-Risk Jurisdictions subject to a Call for Action, and Jurisdictions under Increased Monitoring (statements of 19 June 2026)"),
 (N,     "FCA Final Notices: Monzo Bank Ltd (July 2025); Barclays Bank UK plc (July 2025); Nationwide Building Society (December 2025)"),
 (N,     ""),
 (MU,    "Synthetic and illustrative only. Northgate Bank UK Limited does not exist and no real customer data appears in this workbook. "
         "This is a self-directed learning project and is not professional compliance advice."),
]
r = 2
for font, text in rows:
    c = ws.cell(row=r, column=2, value=text)
    c.font = font
    c.alignment = WRAP
    if font is TITLE:
        ws.row_dimensions[r].height = 22
    elif len(text) > 110:
        ws.row_dimensions[r].height = 15 * (len(text) // 110 + 1)
    r += 1
ws.freeze_panes = "A3"

# ----------------------------------------------------------------- Categories
ws = wb.create_sheet("Categories")
header(ws, 1, ["Code", "Category", "Category weight", "Factors", "Factor weights sum", "Check"],
       [8, 24, 16, 10, 18, 10])
for i, (code, name, w) in enumerate(CATEGORIES):
    r = 2 + i
    ws.cell(row=r, column=1, value=code).font = B
    ws.cell(row=r, column=2, value=name).font = N
    c = ws.cell(row=r, column=3, value=w)
    c.font, c.fill, c.number_format, c.alignment = IN, IN_FILL, "0%", CTR
    c = ws.cell(row=r, column=4, value="=COUNTIF(Factors!$B$%d:$B$%d,$A%d)" % (FR[0], FR[1], r))
    c.font, c.alignment = N, CTR
    c = ws.cell(row=r, column=5, value="=SUMIF(Factors!$B$%d:$B$%d,$A%d,Factors!$D$%d:$D$%d)" % (FR[0], FR[1], r, FR[0], FR[1]))
    c.font, c.number_format, c.alignment = N, "0%", CTR
    c = ws.cell(row=r, column=6, value='=IF(ROUND($E%d,6)=1,"OK","CHECK")' % r)
    c.font, c.alignment = B, CTR
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BOX
r = CR[1] + 1
ws.cell(row=r, column=2, value="Total").font = B
c = ws.cell(row=r, column=3, value="=SUM($C$%d:$C$%d)" % CR)
c.font, c.number_format, c.alignment, c.fill = B, "0%", CTR, SUB_FILL
c = ws.cell(row=r, column=6, value='=IF(ROUND($C%d,6)=1,"OK","CHECK")' % r)
c.font, c.alignment = B, CTR
ws.cell(row=r + 2, column=2, value="Weights are judgement, not derived from loss data. Step 5 tests their sensitivity.").font = MU
ws.freeze_panes = "A2"
add_table(ws, "tblCategories", 1, CR[1], 6)

# ----------------------------------------------------------------- Factors
ws = wb.create_sheet("Factors")
header(ws, 1, ["Factor", "Category", "Factor name", "Weight in category", "Effective weight",
               "Lowest level", "Highest level", "Rationale for the weight", "Source"],
       [9, 10, 34, 13, 13, 11, 11, 74, 46])
for i, (code, cat, name, w, rationale, source) in enumerate(FACTORS):
    r = 2 + i
    ws.cell(row=r, column=1, value=code).font = B
    ws.cell(row=r, column=2, value=cat).font = N
    ws.cell(row=r, column=2).alignment = CTR
    ws.cell(row=r, column=3, value=name).font = N
    ws.cell(row=r, column=3).alignment = WRAP
    c = ws.cell(row=r, column=4, value=w)
    c.font, c.fill, c.number_format, c.alignment = IN, IN_FILL, "0%", CTR
    c = ws.cell(row=r, column=5,
        value="=INDEX(Categories!$C$%d:$C$%d,MATCH($B%d,Categories!$A$%d:$A$%d,0))*$D%d" % (CR[0], CR[1], r, CR[0], CR[1], r))
    c.font, c.number_format, c.alignment = N, "0.00%", CTR
    c = ws.cell(row=r, column=6,
        value="=SUMPRODUCT(MIN((Levels!$A$%d:$A$%d=$A%d)*Levels!$C$%d:$C$%d+(Levels!$A$%d:$A$%d<>$A%d)*99))"
              % (LR[0], LR[1], r, LR[0], LR[1], LR[0], LR[1], r))
    c.font, c.alignment = N, CTR
    c = ws.cell(row=r, column=7,
        value="=SUMPRODUCT(MAX((Levels!$A$%d:$A$%d=$A%d)*Levels!$C$%d:$C$%d))" % (LR[0], LR[1], r, LR[0], LR[1]))
    c.font, c.alignment = N, CTR
    c = ws.cell(row=r, column=8, value=rationale); c.font, c.alignment = N, WRAP
    c = ws.cell(row=r, column=9, value=source);    c.font, c.alignment = N, WRAP
    ws.row_dimensions[r].height = 15 * (len(rationale) // 74 + 1)
r = FR[1] + 1
ws.cell(row=r, column=3, value="Total effective weight").font = B
c = ws.cell(row=r, column=5, value="=SUM($E$%d:$E$%d)" % FR)
c.font, c.number_format, c.alignment, c.fill = B, "0.00%", CTR, SUB_FILL
ws.freeze_panes = "D2"
add_table(ws, "tblFactors", 1, FR[1], 9)

# ----------------------------------------------------------------- Levels
ws = wb.create_sheet("Levels")
header(ws, 1, ["Factor", "Factor name", "Score", "What this level means", "Mandatory escalator"],
       [9, 34, 8, 96, 34])
for i, (code, score, definition, esc) in enumerate(LEVELS):
    r = 2 + i
    ws.cell(row=r, column=1, value=code).font = B
    c = ws.cell(row=r, column=2,
        value="=INDEX(Factors!$C$%d:$C$%d,MATCH($A%d,Factors!$A$%d:$A$%d,0))" % (FR[0], FR[1], r, FR[0], FR[1]))
    c.font, c.alignment = N, WRAP
    c = ws.cell(row=r, column=3, value=score); c.font, c.alignment = N, CTR
    c = ws.cell(row=r, column=4, value=definition); c.font, c.alignment = N, WRAP
    c = ws.cell(row=r, column=5, value=esc); c.font, c.alignment = N, WRAP
    ws.row_dimensions[r].height = 15 * (len(definition) // 96 + 1)
ws.freeze_panes = "A2"
add_table(ws, "tblLevels", 1, LR[1], 5)

# ----------------------------------------------------------------- Escalators
ws = wb.create_sheet("Escalators")
header(ws, 1, ["Reference", "Effect", "Condition"], [12, 16, 116])
for i, (ref, typ, desc) in enumerate(ESCALATORS):
    r = 2 + i
    ws.cell(row=r, column=1, value=ref).font = B
    ws.cell(row=r, column=2, value=typ).font = N
    c = ws.cell(row=r, column=3, value=desc); c.font, c.alignment = N, WRAP
    ws.row_dimensions[r].height = 15 * (len(desc) // 116 + 1)
r = len(ESCALATORS) + 3
ws.cell(row=r, column=3, value="Mandatory High overrides the arithmetic. A weighted average dilutes a single severe factor by design; "
                               "this list is the compensating control. Prohibited relationships are not scored at all.").font = MU
ws.cell(row=r, column=3).alignment = WRAP
ws.freeze_panes = "A2"
add_table(ws, "tblEscalators", 1, len(ESCALATORS) + 1, 3)

# ----------------------------------------------------------------- Example
ws = wb.create_sheet("Example")
ws.sheet_view.showGridLines = False
c = ws.cell(row=1, column=1, value="Worked example: customer 4417")
c.font = TITLE
c = ws.cell(row=2, column=1, value="Sole trader, UK resident, onboarded through the mobile app with electronic identity verification. "
    "Trades in used motor vehicles. Declares turnover of GBP 15,000 to GBP 25,000 a month, of which roughly half is expected in cash. "
    "No adverse media, no PEP match, simple ownership. Scores in blue are the analyst's selections; everything else is calculated.")
c.font, c.alignment = MU, WRAP
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 46

header(ws, 4, ["Factor", "Factor name", "Score", "Effective weight", "Contribution", "Level selected"],
       [9, 34, 8, 14, 13, 92])
E0 = 5
for i, (code, cat, name, w, rationale, source) in enumerate(FACTORS):
    r = E0 + i
    ws.cell(row=r, column=1, value=code).font = B
    c = ws.cell(row=r, column=2,
        value="=INDEX(Factors!$C$%d:$C$%d,MATCH($A%d,Factors!$A$%d:$A$%d,0))" % (FR[0], FR[1], r, FR[0], FR[1]))
    c.font = N
    c = ws.cell(row=r, column=3, value=EXAMPLE[code])
    c.font, c.fill, c.alignment = IN, IN_FILL, CTR
    c = ws.cell(row=r, column=4,
        value="=INDEX(Factors!$E$%d:$E$%d,MATCH($A%d,Factors!$A$%d:$A$%d,0))" % (FR[0], FR[1], r, FR[0], FR[1]))
    c.font, c.number_format, c.alignment = N, "0.00%", CTR
    c = ws.cell(row=r, column=5, value="=$D%d*$C%d" % (r, r))
    c.font, c.number_format, c.alignment = N, "0.000", CTR
    c = ws.cell(row=r, column=6,
        value="=INDEX(Levels!$D$%d:$D$%d,MATCH(1,INDEX((Levels!$A$%d:$A$%d=$A%d)*(Levels!$C$%d:$C$%d=$C%d),0),0))"
              % (LR[0], LR[1], LR[0], LR[1], r, LR[0], LR[1], r))
    c.font, c.alignment = N, WRAP
    ws.row_dimensions[r].height = 28
E1 = E0 + NF - 1

# category subtotals
header(ws, 4, ["Factor", "Factor name", "Score", "Effective weight", "Contribution", "Level selected"],
       [9, 34, 8, 14, 13, 92])
sc = 8
for j, (code, name, w) in enumerate(CATEGORIES):
    r = 4 + j
    ws.cell(row=r, column=sc, value=code).font = B
    ws.cell(row=r, column=sc + 1, value=name).font = N
    c = ws.cell(row=r, column=sc + 2,
        value="=SUMPRODUCT((Factors!$B$%d:$B$%d=$%s%d)*Factors!$D$%d:$D$%d*$C$%d:$C$%d)"
              % (FR[0], FR[1], get_column_letter(sc), r, FR[0], FR[1], E0, E1))
    c.font, c.number_format, c.alignment = N, "0.00", CTR
    c = ws.cell(row=r, column=sc + 3,
        value="=INDEX(Categories!$C$%d:$C$%d,MATCH($%s%d,Categories!$A$%d:$A$%d,0))"
              % (CR[0], CR[1], get_column_letter(sc), r, CR[0], CR[1]))
    c.font, c.number_format, c.alignment = N, "0%", CTR
ws.cell(row=3, column=sc, value="Category scores").font = B
for k, (lbl, wdt) in enumerate([("Code", 9), ("Category", 22), ("Score", 9), ("Weight", 9)]):
    cc = ws.cell(row=4, column=sc + k, value=lbl)
    cc.font, cc.fill, cc.alignment = H, HDR_FILL, CTR
    ws.column_dimensions[get_column_letter(sc + k)].width = wdt
for j in range(len(CATEGORIES)):
    for k in range(4):
        ws.cell(row=5 + j, column=sc + k).border = BOX
# shift category rows down by one so they start under the header at row 4
for j, (code, name, w) in enumerate(CATEGORIES):
    src, dst = 4 + j, 5 + j
    ws.cell(row=dst, column=sc, value=code).font = B
    ws.cell(row=dst, column=sc + 1, value=name).font = N
    c = ws.cell(row=dst, column=sc + 2,
        value="=SUMPRODUCT((Factors!$B$%d:$B$%d=$%s%d)*Factors!$D$%d:$D$%d*$C$%d:$C$%d)"
              % (FR[0], FR[1], get_column_letter(sc), dst, FR[0], FR[1], E0, E1))
    c.font, c.number_format, c.alignment = N, "0.00", CTR
    c = ws.cell(row=dst, column=sc + 3,
        value="=INDEX(Categories!$C$%d:$C$%d,MATCH($%s%d,Categories!$A$%d:$A$%d,0))"
              % (CR[0], CR[1], get_column_letter(sc), dst, CR[0], CR[1]))
    c.font, c.number_format, c.alignment = N, "0%", CTR

# result block
res = E1 + 2
labels = [
 ("Weighted score", "=SUM($E$%d:$E$%d)" % (E0, E1), "0.0000"),
 ("Rounded to two decimal places", "=ROUND($I%d,2)" % res, "0.00"),
 ("Band from the arithmetic", '=IF($I%d<=2,"Low",IF($I%d<=3.49,"Medium","High"))' % (res + 1, res + 1), None),
 ("Mandatory escalator applies (Yes or No)", "No", None),
 ("Final customer risk rating", '=IF($I%d="Yes","High",$I%d)' % (res + 3, res + 2), None),
]
for k, (lbl, formula, fmt) in enumerate(labels):
    r = res + k
    c = ws.cell(row=r, column=7, value=lbl); c.font = B
    c = ws.cell(row=r, column=9, value=formula)
    c.alignment = CTR
    if lbl.startswith("Mandatory"):
        c.font, c.fill = IN, IN_FILL
    elif lbl.startswith("Final"):
        c.font, c.fill = B, OK_FILL
    else:
        c.font = N
    if fmt:
        c.number_format = fmt
ws.cell(row=res + len(labels) + 1, column=7,
        value="The model rates this customer just inside Medium. It is cash-intensive, non-face-to-face and in a sector the "
              "National Risk Assessment treats as vulnerable, yet the geography category scores the minimum and pulls the "
              "average down. Whether that is the right answer is what Step 4 is for.").font = MU
ws.cell(row=res + len(labels) + 1, column=7).alignment = WRAP
ws.merge_cells(start_row=res + len(labels) + 1, start_column=7, end_row=res + len(labels) + 3, end_column=11)
ws.freeze_panes = "A5"

# ----------------------------------------------------------------- Checks
ws = wb.create_sheet("Checks")
ws.sheet_view.showGridLines = False
c = ws.cell(row=1, column=1, value="Integrity checks"); c.font = TITLE
ws.cell(row=2, column=1, value="Every row must read OK. These recalculate whenever a weight or a level is changed.").font = MU
header(ws, 4, ["#", "What is tested", "Result", "Expected", "Status"], [5, 68, 12, 12, 12])
CHECKS = [
 ("Category weights sum to 100%",                      "=ROUND(SUM(Categories!$C$%d:$C$%d),6)" % CR, 1),
 ("Factor weights sum to 100% inside every category",  "=SUMPRODUCT(--(ROUND(Categories!$E$%d:$E$%d,6)<>1))" % CR, 0),
 ("Effective weights sum to 100%",                     "=ROUND(SUM(Factors!$E$%d:$E$%d),6)" % FR, 1),
 ("No duplicate factor codes",                         "=SUMPRODUCT(--(COUNTIF(Factors!$A$%d:$A$%d,Factors!$A$%d:$A$%d)>1))" % (FR[0], FR[1], FR[0], FR[1]), 0),
 ("Every level row points at a factor that exists",    "=SUMPRODUCT(--(COUNTIF(Factors!$A$%d:$A$%d,Levels!$A$%d:$A$%d)=0))" % (FR[0], FR[1], LR[0], LR[1]), 0),
 ("Every factor has at least four levels",             "=SUMPRODUCT(--(COUNTIF(Levels!$A$%d:$A$%d,Factors!$A$%d:$A$%d)<4))" % (LR[0], LR[1], FR[0], FR[1]), 0),
 ("No score outside the range 1 to 5",                 "=SUMPRODUCT(--((Levels!$C$%d:$C$%d<1)+(Levels!$C$%d:$C$%d>5)>0))" % (LR[0], LR[1], LR[0], LR[1]), 0),
 ("No duplicate score within a factor",                "=SUMPRODUCT(--(COUNTIFS(Levels!$A$%d:$A$%d,Levels!$A$%d:$A$%d,Levels!$C$%d:$C$%d,Levels!$C$%d:$C$%d)>1))" % (LR[0], LR[1], LR[0], LR[1], LR[0], LR[1], LR[0], LR[1]), 0),
 ("Every factor's lowest level is 1",                  "=SUMPRODUCT(--(Factors!$F$%d:$F$%d<>1))" % FR, 0),
 ("Every factor's highest level is 5",                 "=SUMPRODUCT(--(Factors!$G$%d:$G$%d<>5))" % FR, 0),
 ("Lowest overall score the model can produce",        "=ROUND(SUMPRODUCT(Factors!$E$%d:$E$%d,Factors!$F$%d:$F$%d),2)" % (FR[0], FR[1], FR[0], FR[1]), 1),
 ("Highest overall score the model can produce",       "=ROUND(SUMPRODUCT(Factors!$E$%d:$E$%d,Factors!$G$%d:$G$%d),2)" % (FR[0], FR[1], FR[0], FR[1]), 5),
 ("Worked example scores exactly as documented (2.07)","=ROUND(SUM(Example!$E$%d:$E$%d),2)" % (E0, E1), 2.07),
]
for i, (label, formula, expected) in enumerate(CHECKS):
    r = 5 + i
    c = ws.cell(row=r, column=1, value=i + 1); c.font, c.alignment = N, CTR
    c = ws.cell(row=r, column=2, value=label); c.font, c.alignment = N, WRAP
    c = ws.cell(row=r, column=3, value=formula); c.font, c.alignment = N, CTR
    c.number_format = "0.00" if isinstance(expected, float) else "General"
    c = ws.cell(row=r, column=4, value=expected); c.font, c.alignment = N, CTR
    c.number_format = "0.00" if isinstance(expected, float) else "General"
    c = ws.cell(row=r, column=5, value='=IF($C%d=$D%d,"OK","CHECK")' % (r, r))
    c.font, c.alignment, c.fill = B, CTR, OK_FILL
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = BOX
r = 5 + len(CHECKS) + 1
ws.cell(row=r, column=2, value="Overall status").font = B
c = ws.cell(row=r, column=5, value='=IF(COUNTIF($E$5:$E$%d,"CHECK")=0,"ALL OK","FIX")' % (4 + len(CHECKS)))
c.font, c.alignment, c.fill = B, CTR, OK_FILL
ws.cell(row=r + 2, column=2, value="Checks 11 and 12 cannot fail while checks 1 to 3, 9 and 10 pass: if every factor runs from 1 to 5 "
    "and the weights sum to 100%, the achievable range is 1.00 to 5.00 by arithmetic. They are kept because they evaluate the aggregation "
    "formula itself, so a mistyped range or a transposed weight vector fails here and nowhere else in this workbook. They say nothing "
    "about the level definitions. See docs/02-risk-factor-rationale.md, 8.2a.").font = MU
ws.cell(row=r + 2, column=2).alignment = WRAP
ws.merge_cells(start_row=r + 2, start_column=2, end_row=r + 3, end_column=5)
ws.freeze_panes = "A5"

for s in wb.worksheets:
    s.sheet_properties.tabColor = "14454E"

out = REPO + "/model/risk-factor-library.xlsx"
wb.save(out)
print("saved", out)
