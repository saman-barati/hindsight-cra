# -*- coding: utf-8 -*-
"""Write backtest/README.md and the six case files, pulling every score straight
from the workbook so no figure in the prose can drift from the model.

v0.2 (post-review). Each case is now reported as a range rather than a point, the
ceiling of each reconstruction is stated before any conclusion is drawn from it, and
the penalty total is attributed to the notices rather than to the six customers.
"""
import os
from paths import REPO
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
from backtest_block import BACKTEST, CASES
from data import FACTORS

XL = REPO + "/model/customer-risk-model.xlsx"
BT = REPO + "/backtest"
wb = load_workbook(XL, data_only=True)
ws, wsw = wb["Backtest"], wb["Weights"]

FO = [f[0] for f in FACTORS]
FNAME = {f[0]: f[2] for f in FACTORS}
WEFF = {wsw.cell(r, 1).value: wsw.cell(r, 3).value for r in range(2, 22)}

R = {}
for r in range(4, 4 + len(BACKTEST)):
    ref = ws.cell(r, 1).value
    R[ref] = {
        "name": ws.cell(r, 2).value,
        "levels": {FO[i]: ws.cell(r, 3 + i).value for i in range(20)},
        "scores": {FO[i]: ws.cell(r, 23 + i).value for i in range(20)},
        "w": ws.cell(r, 43).value, "x": ws.cell(r, 44).value,
        "band": ws.cell(r, 45).value, "esc": ws.cell(r, 46).value, "final": ws.cell(r, 47).value,
        "C": ws.cell(r, 48).value, "G": ws.cell(r, 49).value, "P": ws.cell(r, 50).value,
        "D": ws.cell(r, 51).value, "A": ws.cell(r, 52).value,
        "ruleB": ws.cell(r, 53).value, "rec": ws.cell(r, 54).value,
    }

SLUG = {"FO": "natwest-fowler-oldfield", "SAN": "santander-translations-company",
        "STU": "barclays-stunt-and-co", "WTK": "barclays-wealthtek",
        "MON": "monzo-landmark-address", "NAT": "nationwide-business-on-personal"}
SHORT = {"FO": "NatWest, £264.8m, criminal fine, Dec 2021",
         "SAN": "Santander UK Plc, £107.8m, Dec 2022",
         "STU": "Barclays Bank Plc, £39.3m, Jul 2025",
         "WTK": "Barclays Bank UK Plc, £3.1m, Jul 2025",
         "MON": "Monzo Bank Ltd, £21.1m, Jul 2025",
         "NAT": "Nationwide Building Society, £44.1m, Dec 2025"}
TOTAL = "£480 million"
ORDER = ["Low", "Medium", "High"]
# The cases whose pair really is a favourable/adverse bracket. STU is a pair of dates and SAN a pair of
# states of knowledge, so neither belongs in a statistic about how far a reading can move a score.
BRACKETED = ["FO", "WTK", "MON", "NAT"]


def pinned(refs):
    ps = [f for f in FO if all(R[r]["scores"][f] == 1 for r in refs)]
    wp = sum(WEFF[f] for f in ps)
    return ps, wp, 1 * wp + 5 * (1 - wp)


def rng(refs, key="x"):
    v = [R[r][key] for r in refs]
    return min(v), max(v)


def finals(refs):
    return sorted(set(R[r]["final"] for r in refs), key=ORDER.index)


def fstr(refs):
    f = finals(refs)
    return f[0] if len(f) == 1 else "%s to %s" % (f[0], f[-1])


os.makedirs(BT + "/cases", exist_ok=True)


def w(path, text):
    open(path, "w", encoding="utf-8").write(text.strip() + "\n")


HEAD = """**Document reference:** HND-CRA-007
**Version:** 0.3
**Author:** Saman Barati
**Date:** September 2026"""

FOOT = """---

*Reconstruction, not a finding.* This case file rebuilds a customer from what a published FCA notice
records. It is not a claim about what the firm's file actually contained, and it is not a criticism of
any individual. The purpose is to test a model, not to re-try a case."""


def diff_table(a, b, names):
    L = ["| Factor | %s | %s |" % names, "|---|---|---|"]
    for f in FO:
        x, y = R[a]["levels"][f], R[b]["levels"][f]
        if x != y:
            L.append("| %s %s | %s (%d) | **%s (%d)** |"
                     % (f, FNAME[f], x, R[a]["scores"][f], y, R[b]["scores"][f]))
    return "\n".join(L)


def result_table(refs, names):
    L = ["| | " + " | ".join(names) + " |", "|---|" + "---|" * len(refs)]
    for label, key, fmt in [("Weighted score", "w", "%.4f"), ("Rounded", "x", "%.2f"),
                            ("Band from the arithmetic", "band", "%s"),
                            ("Mandatory escalator", "esc", "%s"),
                            ("**Final rating**", "final", "**%s**")]:
        L.append("| %s | " % label + " | ".join(fmt % R[r][key] for r in refs) + " |")
    return "\n".join(L)


def ceiling_para(code, refs):
    ps, wp, ceil = pinned(refs)
    return ("%d of the 20 factors are settled at level 1 by the notice itself and take that value in "
            "every reading. They carry **%.2f%% of the model's weight**. Even if every one of the "
            "remaining %d factors were pushed to 5, this reconstruction could not score above **%.2f**, "
            "and the High band begins at 3.50. Nothing this case file says about the arithmetic is a "
            "discovery; the ceiling was fixed before any judgement was made."
            % (len(ps), 100 * wp, 20 - len(ps), ceil))


# ------------------------------------------------------------------ README
case_rows = []
for code, name, refs in CASES:
    lo, hi = rng(refs)
    _, _, ceil = pinned(refs)
    case_rows.append("| [%s](cases/%s.md) | %s | %.2f – %.2f | %.2f | %s |"
                     % (code, SLUG[code], SHORT[code], lo, hi, ceil, fstr(refs)))

hi_cases = [c for c, n, refs in CASES if "High" in finals(refs)]
arith_high = [ref for ref in R if R[ref]["band"] == "High"]
maxceil = max(pinned(refs)[2] for c, n, refs in CASES)
# The share of the model's weight the notices pin at level 1, across the six cases. Computed, not
# typed: an earlier version hard-coded this range in six documents and one of them went stale the
# moment a single case was rebuilt.
PINS = [100 * pinned(refs)[1] for c, n, refs in CASES]
PINLO, PINHI = min(PINS), max(PINS)

w(BT + "/README.md", """
# Step 4 — the back-test

%s

## The question

Every risk model can be checked for internal consistency. Very few are checked against customers who
are already known to have caused harm. This step takes six customers named in published FCA
enforcement notices, rebuilds each of them as they looked **on the day the bank took them on**, and
runs them through the same library, weights and rules as the 400 synthetic customers.

The six cases carry roughly %s of penalties between them. Three things about that number, because it
is the one most likely to be quoted back.

It is arithmetic on the published figures and nothing more: £264,772,619.95 + £107,793,300 +
£44,078,500 + £39,314,700 + £21,091,300 + £3,093,600 = £480,144,019.95.

Those penalties were imposed for control failures across whole customer books over periods of years —
not for these six accounts, and in no case was a single customer the cause of a fine. The customers
are used here because the published material describes them in enough detail to rebuild, which is a
different thing from blaming them for the total.

And the total mixes two kinds of thing. Five are regulatory penalties imposed by the FCA under section
206 of FSMA and set out in final notices. The NatWest figure is a **criminal fine imposed by Southwark
Crown Court** following an FCA prosecution under the Money Laundering Regulations 2007; there is no
final notice for it. Adding them is fair as a measure of scale and wrong as a measure of anything else.

## Ranges, not point estimates

A reconstruction from a published notice involves twenty judgement calls, and the notice settles far
fewer than twenty of them. The first version of this back-test made those calls once and reported a
single score, which gave the reconstruction an authority it had not earned. Every case is now scored
at least twice:

- **the favourable reading** — every factor the notice leaves open, read the way most favourable to the customer;
- **the adverse reading** — every open factor at the worst level a competent analyst could have defended on the same information.

Neither is the answer. The pair is, and the width of the pair is the honest measure of how much of the
result is reconstruction rather than fact. Where the notice settles a factor, both readings carry the
same level.

**Two of the six are labelled `-a` and `-b` rather than `-lo` and `-hi`, because their pair is not a
favourable/adverse bracket at all.** Stunt & Co is a pair of **dates**, eleven days apart. Santander is
a pair of **states of knowledge**: the file as the bank recorded it, and the same file with the one
verification the FCA found was missing. Anywhere this repository describes the method as
favourable-and-adverse without that qualification, it is describing four of the six cases.

## The six cases

| Case | Notice | Score range | Ceiling | Final rating |
|---|---|---|---|---|
%s

"Ceiling" is the highest score this reconstruction could reach if every factor the notice does not
settle were pushed to 5.

## What the back-test can and cannot show

**It cannot show that these customers should have scored High on the arithmetic.** Read the ceiling
column. The highest ceiling of any of the six is **%.2f**, and the High band begins at 3.50. The
facts the notices settle pin between %.0f%% and %.0f%% of the model's weight at level 1 before a single
judgement is made: for the four business customers, a UK company with UK owners onboarded in branch on
documents seen face to face; for Monzo and Nationwide, an ordinary UK personal customer with no
international activity, which pins even more. A test
whose answer is fixed by its inputs is not a test. **The finding that no reconstruction reaches the
High band on the arithmetic was guaranteed before the exercise began**, and every conclusion below is
written in that knowledge.

**It can show three things the ceiling does not decide.**

1. Whether a **mandatory escalator** fires. That is a yes/no condition on individual levels and is not
   subject to the ceiling at all.
2. How far apart the favourable and adverse readings sit, which measures how much the model's answer
   depends on the analyst rather than on the file.
3. Whether the **rating changes as the file changes** — which is what the Stunt & Co dates and the
   Santander verification both test, and it is where the two most useful findings came from.

## What it found

### 1. Three of the six are caught, and every one of them by the same escalator

%s of the six cases reach High in at least one reading, and all three do it through escalator 5.3(c) —
money service business, trust or company service provider, or dealer in high-value goods — rather than
through a score. **Not one reconstruction, in any reading, reaches the High band on the arithmetic**,
though for the reasons above none of them could have.

One escalator carrying every catch in the back-test is not a comfortable result. It says the model has
one working control at the top of the scale and nineteen factors that mostly decide nothing, and it
puts an uncomfortable amount of weight on the wording of a single level definition. That wording is
the subject of findings 2 and 4.

Stunt & Co scores %.2f on the application file of 16 January 2015 and is rated High from that day,
because the application records the business as gold refining and trading. That is the same level
definition the project had already recommended narrowing, on the ground that it is written as a
description of a trade rather than as a registration status. **The recommendation would have removed
the control that catches this customer.** It stays in Step 5 as a recommendation, now with that
argument recorded against it.

### 2. One word still decides the most serious case

`FO-lo` and `FO-mid` are the same twenty-factor file. They differ in one reading of one factor: whether
a Bradford jewellery business is a *cash-intensive trade* or a *dealer in high-value goods*.

%s

C3 carries 7.5%% of the model, so one level of difference moves the score by exactly 0.075 — from
%.2f to %.2f. Neither crosses a band boundary. The rating changes anyway, because level 5 fires
escalator 5.3(c) and level 4 does not. The score is not what decides this case; a definition is.

### 3. The declared profile is the model's blind spot, in both directions

Fowler Oldfield declared no cash and the model scored what was declared: A2 stays at level 1 in every
reading, because nothing at onboarding contradicted it. Around £264 million in cash followed. A model
that scores declarations cannot catch a customer who lies at onboarding, and no weighting fixes that.

Stunt & Co is the mirror image. Its application form carried an anticipated turnover of £500,000
manually amended to £3 million, with no recorded explanation — an inconsistency **on the face of the
file on day one**. The model has a factor for exactly this (A5, plausibility) and gives it 1.5%% of the
weight. It moves the score by 0.045.

### 4. The model can be right and still be wrong, because it scores what it is given

Santander's customer is the sharpest thing in this back-test.

The application form said "Translation service". Scored from that form the customer is **Low**. The
notice records something the form did not: the customer operated a **money service business**, and
Santander did not identify that at onboarding. Verified, the same customer is **High** on day one,
through the same escalator as findings 1 and 2.

So the model contained the control the whole time. What it did not contain — what no model contains —
is the customer due diligence that fills in the input. Five documents in this project argue about
weights, boundaries and aggregation rules. None of that argument survives contact with a file where
the trade has not been verified.

That is recorded at methodology 11.15, and it is the finding I would keep if I could only keep one.

### 5. Nationwide's customer really was ordinary

Not every case is a failure of the rating. Nationwide's personal customer looks ordinary on day one in
both readings and is rated Low, correctly: nothing in a customer risk assessment could have predicted a
fraud that had not happened yet, against a scheme that did not yet exist. What failed there was the
refresh and the monitoring, which is what the event-driven review triggers at methodology 7.2 are for —
including trigger (d), which exists in this project because of that notice.

### 6. The width of the range is itself a result

Across the four cases scored as a favourable/adverse pair, the gap runs from %.2f to %.2f points. On a
scale where the whole synthetic population spans 1.19 to 2.76, a spread of that size means the
analyst's reading of an incomplete file moves a customer about as far as most real differences between
customers do. The two labelled pairs are excluded from that figure, because their two rows differ by a
date and by a verification rather than by a judgement.

## Where the reconstructions come from

Every level recorded in the `Backtest` sheet traces to one of these, and nothing else:

| Case | Primary source |
|---|---|
| Fowler Oldfield | [FCA press release](https://www.fca.org.uk/news/press-releases/natwest-fined-264.8million-anti-money-laundering-failures) and the [agreed statement of facts](https://www.fca.org.uk/publication/corporate/agreed-statement-facts-fca-national-westminster-bank.pdf) |
| Santander | [Final notice, 8 December 2022](https://www.fca.org.uk/publication/final-notices/santander-uk-plc-2022.pdf) |
| Stunt & Co | [Barclays Bank Plc final notice, 14 July 2025](https://www.fca.org.uk/publication/final-notices/barclays-bank-plc-2025.pdf) |
| WealthTek | [Barclays Bank UK Plc final notice, 14 July 2025](https://www.fca.org.uk/publication/final-notices/barclays-bank-uk-plc-2025.pdf) |
| Monzo | [Final notice, 7 July 2025](https://www.fca.org.uk/publication/final-notices/monzo-bank-limited.pdf) |
| Nationwide | [Final notice, 11 December 2025](https://www.fca.org.uk/publication/final-notices/nationwide-building-society-2025.pdf) |

Note that the two Barclays matters are **two entities and two separate final notices**: Barclays Bank
Plc for Stunt & Co, Barclays Bank UK Plc for WealthTek. They are announced together in one press
release, which is where the "£42 million" headline figure comes from. One notice is not authority for
the other, and the two are cited separately throughout.

## Reading the workbook

The `Backtest` sheet of `model/customer-risk-model.xlsx` carries all %d rows: the recorded level for
each of the twenty factors, the score each resolves to, the weighted result, the band, the escalator
test and the final rating. The four integrity checks covering the back-test are on the `Checks` sheet.

## The case files

%s

---

*Reconstructions, not findings.* These files rebuild customers from published FCA notices. They are
not claims about what any firm's file actually contained, and not criticism of any individual.
""" % (
    HEAD, TOTAL, "\n".join(case_rows), maxceil, PINLO, PINHI,
    ["None", "One", "Two", "Three", "Four", "Five", "Six"][len(hi_cases)],
    R["STU-a"]["x"],
    diff_table("FO-lo", "FO-mid", ("Cash-intensive reading", "High-value dealer reading")),
    R["FO-lo"]["x"], R["FO-mid"]["x"],
    min(rng(refs)[1] - rng(refs)[0] for c, n, refs in CASES if c in BRACKETED),
    max(rng(refs)[1] - rng(refs)[0] for c, n, refs in CASES if c in BRACKETED),
    len(BACKTEST),
    "\n".join("- [%s](cases/%s.md) — %s" % (n, SLUG[c], SHORT[c]) for c, n, refs in CASES),
))


# ------------------------------------------------------------------ case files
def case(code, title, meta, story, settled, readings_note, teaches, extra=""):
    refs = dict((c[0], c[2]) for c in CASES)[code]
    names = {"FO": ("Cash-intensive reading", "High-value dealer reading", "Adverse reading"),
             "STU": ("16 January 2015", "27 January 2015"),
             "SAN": ("The file as recorded", "With the business verified")}.get(
        code, ("Favourable reading", "Adverse reading"))
    lo, hi = refs[0], refs[-1]
    body = """
# %s

%s

%s

## What happened

%s

## What the notice settles

%s

## Where the readings differ

%s

%s

## What the model says

%s

%s

## What this teaches the model

%s

%s

%s
""" % (title, HEAD, meta, story, settled, readings_note,
       diff_table(lo, hi, names[:2]) if len(refs) == 2 else diff_table(refs[0], refs[-1], (names[0], names[-1])),
       result_table(refs, list(names[:len(refs)])),
       ceiling_para(code, refs), teaches.strip(), extra, FOOT)
    while "\n\n\n" in body:
        body = body.replace("\n\n\n", "\n\n")
    w(BT + "/cases/%s.md" % SLUG[code], body)


case("FO", "NatWest and Fowler Oldfield", """
| | |
|---|---|
| **Firm** | National Westminster Bank Plc |
| **Penalty** | £264,772,619.95 |
| **Date** | Sentenced 13 December 2021, Southwark Crown Court, before Mrs Justice Cockerill |
| **Basis** | Criminal conviction on FCA prosecution, following a guilty plea on 7 October 2021 to offences under regulations 8(1), 8(3) and 14(1) of the Money Laundering Regulations 2007. **This is a criminal fine, not an FCA final notice** — the only one of the six that is. |
| **Relevant period** | 8 November 2012 to 23 June 2016, the outer range across the three offences |
| **Sources** | [FCA press release](https://www.fca.org.uk/news/press-releases/natwest-fined-264.8million-anti-money-laundering-failures) &middot; [Agreed statement of facts](https://www.fca.org.uk/publication/corporate/agreed-statement-facts-fca-national-westminster-bank.pdf) |
""".strip(), """
The FCA describes Fowler Oldfield as "a jewellery business based in Bradford". At the start of the
relationship **NatWest understood that it would not handle cash from the business** — the relationship
manager's note reads "We will not handle any cash for this business" (agreed statement of facts, 80).
Around **£365 million was subsequently deposited with the bank**, of which roughly **£264 million was
in cash**.

The agreed statement of facts records significant amounts of Scottish banknotes deposited throughout
England; notes carrying a prominent musty smell at one cash centre; cash brought into a branch
uncounted in black bin liners at another; and a system error under which cash deposited through cash
centres was read as cheque deposits and subjected to the less stringent rules that apply to cheques.
Staff raised concerns. Nothing changed. In the FCA's press release the sentencing judge, Mrs Justice
Cockerill, is quoted as finding that "the Bank was functionally vital. Without the Bank's failures -
the money could not be effectively laundered".
""".strip(), """
- A UK company with natural-person owners, onboarded in branch with documents seen face to face.
- The business is described three different ways in the FCA's own material: "a jewellery business
  based in Bradford" in the press release; staff recorded as saying "they are not pawnbrokers or
  jewellers but precious metal dealers" (120); and the bank itself having "erroneously described
  Fowler Oldfield as a pawnbroker" (111). The library reads a jeweller and a precious metal dealer
  differently, and that is finding two below.
- **The file recorded that the account would not handle cash.** This is the fact the case turns on, and
  it is settled: A2 sits at level 1 in every reading. Note who said it. The FCA records this as the
  *bank's* understanding, not as a customer declaration.
""".strip(), """
Three rows are scored. `FO-lo` and `FO-mid` are the **same file** and differ only in how C3 is read —
that comparison is the point of the case, so every other factor is held constant. `FO-hi` then takes
the high-value-dealer reading and pushes every remaining open factor to the worst level the notice
permits.
""".strip(), """
1. **The model scores what is on the file, and cannot ask who put it there.** A2 stays at level 1 in all
   three readings. The FCA records "no cash" as the bank's own understanding rather than as something
   the customer stated, and to this model that distinction does not exist: an assumption written into a
   file is scored exactly like a verified fact. Nothing in twenty factors asks where an answer came
   from. No weighting fixes that. What would is a control that compares what the file says with what
   the account does, which is monitoring, not rating.

2. **One word decides the rating, and the FCA's own material supplies three of them.** C3 carries 7.5%
   of the model, so reading the business as a high-value dealer rather than a cash-intensive trade moves
   the score by 0.075 — not enough to change a band. It changes the rating anyway, because level 5 fires
   escalator 5.3(c) and level 4 does not. The escalator is doing the work; the score is decoration. And
   "jeweller", "pawnbroker" and "precious metal dealer" are not synonyms in this library, which is a
   defect in a scale that asks an analyst to pick one word for a business.

3. **The statutory definition does not support the high-value-dealer reading at onboarding.**
   Regulation 14 of the MLRs ("High value dealers, casinos, auction platforms and art market
   participants") defines a high value dealer as a firm or sole trader trading in goods by way of
   business "when the trader makes or receives, in respect of any transaction, a payment or payments in
   cash of at least £10,000 in total" — a threshold redenominated from 10,000 euros to £10,000 by
   regulation 9 of SI 2026/621 with effect from 30 June 2026. The trigger is cash, and this file
   recorded none. So reading (b) is defensible on the wording *this library* uses and not on the wording
   the regulations use — which sharpens the finding rather than weakening it, because the defect is then
   squarely in the level definition. Step 5 recommends tying the level to registration. The Stunt & Co
   case file records the argument against doing so.
""".strip())

case("SAN", "Santander UK and the translations company", """
| | |
|---|---|
| **Firm** | Santander UK Plc |
| **Penalty** | £107,793,300 |
| **Date** | Final notice 8 December 2022 |
| **Source** | https://www.fca.org.uk/publication/final-notices/santander-uk-plc-2022.pdf |
""".strip(), """
A business opened a Santander account on an application form describing its trade as a "Translation
service", with estimated annual turnover of **£100,000** and expected monthly deposits of **£5,000**
(4.52). Within six months the account was receiving millions, swiftly transferred on to separate
accounts. The bank's own anti-money laundering team recommended closing the account in March 2014, the
decision to close was taken in April 2015, and the account was not closed until September 2015 —
**eighteen months** after the recommendation.

Two things in the notice matter more than that story.

**The customer operated a money service business, and Santander did not identify it as one at
onboarding.** The FCA's finding is that the firm failed to verify the nature of the business. That is
not a monitoring failure a rating could never have reached. It is a failure of the input the rating
depends on.

**The scale of it.** Roughly £298 million passed through the accounts of six money service business
customers before they were closed, the largest single account accounting for about £269 million. This
case file does **not** assert that this customer is that account. The notice anonymises its customers
and I have not established which letter this one carries, so the £269m figure is context for the
population these findings come from, not a fact about this reconstruction.
""".strip(), """
- A UK company, a business account, an application form reading "Translation service", estimated
  annual turnover of £100,000 and expected monthly deposits of £5,000 (4.52).
- **The customer operated a money service business**, which the notice records was not identified at
  onboarding.
- A1 is scored at level 2, not level 1. £5,000 a month sits exactly on the level 1 / level 2 boundary
  and methodology 4.5 sends a straddling figure to the higher level; £100,000 a year confirms it.
""".strip(), """
This pair is not a favourable and an adverse reading. `SAN-a` is the file as Santander recorded it,
with the trade taken from the application form. `SAN-b` is the same file with the one verification the
notice says was missing: the business identified as a money service business, which is level 5 of
factor C3. Three other factors move with it, because a money service business sending occasional
international payments is a different profile from a translation agency.
""".strip(), """
1. **The model had the control. The file did not have the input.** Verified as a money service
   business the customer is rated **High** on day one, through escalator 5.3(c). Taken from the
   application form at face value it is rated **Low**. The gap between those two rows is not a modelling
   choice or a matter of judgement; it is one verification step that the notice says was not carried
   out.

2. **A risk model is only ever as good as the due diligence feeding it.** This project spends five
   documents arguing about weights, bands and aggregation rules. On a file where the trade has not been
   verified, none of it matters: the model faithfully scores what it is given. Recorded at methodology
   11.15, and it is the single most useful thing the back-test produced.

3. **The second half of the failure stands.** The declared £5,000 a month sits at A1 and nothing in the
   model ever compares it with what the account received. That comparison belongs to transaction
   monitoring and to the event-driven review triggers at methodology 7.2. The eighteen months between an
   internal recommendation to close and the closure belong to neither.
""".strip())

case("STU", "Barclays Bank plc and Stunt & Co", """
| | |
|---|---|
| **Firm** | Barclays Bank Plc |
| **Penalty** | £39,314,700 (reduced from £56,163,900) |
| **Date** | Final notice 14 July 2025, announced 16 July 2025 |
| **Source** | https://www.fca.org.uk/publication/final-notices/barclays-bank-plc-2025.pdf |
""".strip(), """
**Read this first.** Paragraph 2.7 of the Final Notice records that **on 4 March 2025 James Stunt was
acquitted** of money laundering charges in relation to monies received by Stunt & Co from Fowler
Oldfield, "on the basis that he had no knowledge or suspicion that those monies were criminal
property". Paragraph 4.138 records the same. The FCA's findings in this notice are about **Barclays'
controls**, and nothing in this case file should be read as saying anything else.

The notice records that Stunt & Co received **£46.8 million from Fowler Oldfield** in electronic
transfers between July 2015 and August 2016. The FCA found that Barclays did not gather enough
information at the start of the relationship and did not carry out proper ongoing monitoring.

The case is here for one narrow reason: the notice reproduces enough of the account opening file to
score it. A customer risk rating is a statement about the information on a file on a particular day,
made without knowing how anything turns out. This case makes that unusually clear, and the rating
below is a fact about a form, not about a person.
""".strip(), """
From the Final Notice of 14 July 2025:

| Paragraph | What the bank held |
|---|---|
| 4.12 | A single director and sole shareholder; the account opened on **16 January 2015** |
| 4.20(a) | The application recorded the business as **"Gold Refining & Trading"** |
| 4.20(b) | An opening investment of **£1,500,000**, paid in from the shareholder's own accounts |
| 4.20(c) | Anticipated turnover of **£500,000, manually amended to £3 million** |
| 4.20(d) | The company **"would not trade outside of the EU"** |
| 4.24 | At a meeting on **27 January 2015**, eleven days after opening: gold sourced from West Africa, principally Ghana and Burkina Faso, sold to high net worth individuals in the Middle East |
| 4.35 | Barclays **classified the customer as low risk**, with no recorded rationale |
| 4.37 | Adverse media checks were run on 15 January 2015; there is **no evidence the results were reviewed** |
| 2.7, 4.138 | The shareholder was **acquitted** of money laundering charges on 4 March 2025 |
""".strip(), """
The two rows here are not a favourable and an adverse reading. They are the **same file on two dates**,
eleven days apart, and both are things the bank held. Nothing in either row uses information that
emerged later.
""".strip(), """
1. **The application form was enough.** "Gold Refining & Trading" is a dealer in high-value goods on the
   wording of C3 level 5, which fires escalator 5.3(c). The model rates this customer **High on
   16 January 2015**, from the application form alone, with no hindsight and no adverse media. The FCA
   records that Barclays rated it low risk and did not write down why.

2. **The recommendation this project had already made would have removed that.** Step 5 recommends
   tying the high-value-dealer level to *registration* rather than to a description of the trade,
   because the description is what makes the Fowler Oldfield reading ambiguous. Stunt & Co declared no
   cash, so it would probably not have been a registered high value dealer. Applying the recommendation
   would have taken this customer from High to Medium. **A fix that resolves one case breaks another**,
   and the recommendation now carries that argument against it rather than being presented as an
   improvement.

3. **The £500,000 amended to £3 million is the sharpest fact in the notice and the model barely feels
   it.** A six-fold manual amendment with no recorded explanation is an unresolved inconsistency on the
   face of the file on day one. It is recorded at A5, which carries 1.5%% of the model and moves the
   score by 0.045. Step 5 recommends moving A5 onto the escalator list; note that the recommendation as
   drafted catches only level 5, and on 16 January this file is at level 4. The recommendation would
   have caught it on 27 January and not before.

4. **Eleven days is not a rating cycle.** Between the two rows the customer's stated trading model went
   from "not outside the EU" to West Africa and the Middle East. The score moves from %.2f to %.2f and
   the rating does not move at all, because it was already High. Had the escalator not fired, nothing in
   this model would have re-rated the customer on the new information: a customer risk assessment is
   produced at onboarding and refreshed on a cycle, and the control that is supposed to react in
   eleven days is the event-driven review trigger at methodology 7.2, not the rating.

5. **"Screening was run but nobody read it" has no level.** C5 sits at level 1 in both rows because the
   notice does not record what the unreviewed results contained. The library has no level for a check
   that was performed and not reviewed, which is not the same as a clear result. That is a gap in the
   library, recorded at methodology 11.10.
""" % (R["STU-a"]["x"], R["STU-b"]["x"]))

case("WTK", "Barclays Bank UK plc and WealthTek", """
| | |
|---|---|
| **Firm** | Barclays Bank UK Plc — a different legal entity from the Stunt & Co case, and a separate final notice |
| **Penalty** | £3,093,600 (reduced from £4,419,500) |
| **Date** | Final notice 14 July 2025, announced 16 July 2025 |
| **Source** | https://www.fca.org.uk/publication/final-notices/barclays-bank-uk-plc-2025.pdf |
""".strip(), """
Barclays UK opened a client money account for WealthTek without checking the Financial Services
Register. The final notice records that "the Authority's FS Register was not reviewed for any
requirement restricting WealthTek's ability to hold client money" (4.30), and that had it been checked
Barclays UK "would have identified that the Authority had imposed a requirement on its Part 4A
authorisation preventing WealthTek from holding client money" (2.7).

The check that was missing is free, takes under a minute, and is public.
""".strip(), """
- A UK company, a client money account, regular international payments, a large number of individual
  counterparties.
- **The customer was an FCA-authorised investment firm.** The library has no level for that, which is
  finding two below.
""".strip(), """
The notice is about a check that was not performed rather than about the customer's profile, so most
factors are open. The adverse reading pushes ownership, corridor, verification and international share
to the worst defensible level and still moves the score by less than a quarter of a point.
""".strip(), """
1. **The model cannot ask a question the library does not contain.** Nothing in twenty factors asks
   whether the customer holds the permissions its stated business requires. Adding that is trivial —
   it is a free public lookup — and its absence is the whole case. Recorded at methodology 11.11.

2. **The library has no home for a financial-sector customer.** C3 level 2 reads "regulated profession
   with a named supervisory body **outside the financial sector**". WealthTek is a wealth manager. Level
   3 is retail, hospitality, transport and construction; level 4 is cash-intensive trades; level 5 is
   MSBs, TCSPs and high-value dealers. A bank's customer that is itself a regulated financial firm has
   nowhere to sit, and is forced into level 2 in both readings here. That is a defect in the scale, not
   a judgement call, and it is recorded at methodology 11.12.

3. **D3 records the wrong register.** Factor D3 levels 4 and 5 turn on whether an *introducing firm's*
   register entry was checked. WealthTek was the customer, not an introducer, so the factor that
   mentions the Financial Services Register is the one factor that cannot capture this case. The scale
   also mixes two dimensions — how the customer arrived, and whether a check was done — which is why it
   reads as if it covers something it does not.
""".strip())

case("MON", "Monzo and the landmark address", """
| | |
|---|---|
| **Firm** | Monzo Bank Ltd |
| **Penalty** | £21,091,300 (reduced from £30,130,475) |
| **Date** | Final notice 7 July 2025 |
| **Source** | https://www.fca.org.uk/publication/final-notices/monzo-bank-limited.pdf |
""".strip(), """
Monzo onboarded customers on the basis of limited and, in some cases, **obviously implausible
information — including customers who gave well known London landmarks as their home address**. The
final notice names Buckingham Palace and 10 Downing Street (4.50). Its customer risk assessments did
not keep pace as the customer base grew almost tenfold, from around 600,000 in 2018 to over 5.8 million
in 2022. After the FCA restricted it in August 2020 from opening accounts for high-risk customers, it
signed up over 34,000 of them anyway between August 2020 and June 2022.
""".strip(), """
- A personal customer, onboarded digitally through the firm's own app.
- **The stated address is a famous building.** A5 level 5 in the library reads: *the declared profile is
  not credible on the information held, including an address that is a landmark or a non-residential
  building.* That wording was written from this case, in Step 2, before anything was scored. It is
  settled at level 5 in both readings.
""".strip(), """
The notice settles more of this file than any other case — a personal customer has fewer moving parts —
so the range here is narrow. That is itself informative: the result below is not a reconstruction
artefact.
""".strip(), """
1. **The factor that saw the problem was too light to change the answer.** A5 carries 1.5%% of the model.
   Moving it from level 1 to level 5 adds 0.06 to the score. The customer lands comfortably in Low in
   both readings, and the argument defending that weight is still in `docs/02-risk-factor-rationale.md`.
   Both the reasoning and the result stay in the repository: deleting the argument that turned out wrong
   would remove the only evidence that the test worked.

2. **Some facts are not scale-shaped.** "This address is a landmark" is not a degree of risk, it is a
   file that should not have been completed. Regulation 31 already requires a firm to cease transactions
   where it cannot complete customer due diligence. Step 5 recommends moving A5 level 5 onto the
   escalator list, and this case is the clearest argument for it.

3. **The delivery channel category, not the plausibility factor, is what raises this customer at all.**
   The category scores %.2f, against %.2f for expected activity. Digital onboarding with single-source
   electronic verification is doing more work than the landmark address.
""" % (R["MON-lo"]["D"], R["MON-lo"]["A"]))

case("NAT", "Nationwide and the personal account used for business", """
| | |
|---|---|
| **Firm** | Nationwide Building Society |
| **Penalty** | £44,078,500 (reduced from £62,969,297) |
| **Date** | Final notice 11 December 2025, announced 12 December 2025 |
| **Relevant period** | 1 October 2016 to 1 July 2021 |
| **Source** | https://www.fca.org.uk/publication/final-notices/nationwide-building-society-2025.pdf |
""".strip(), """
Nationwide did not keep due diligence and risk assessments up to date across its personal current
account book, and did not adequately monitor transactions. It knew customers were running business
activity through personal accounts in breach of its terms and had no process to manage the financial
crime risk that created.

One customer received **24 fraudulent Covid furlough payments** under the Job Retention Scheme. The
notice records the 24 payments as totalling **£1.35 million over thirteen months**, followed by a
further **£26.01 million over eight days** (2.17). HMRC seized £26.54 million; £820,687 was not
recovered.
""".strip(), """
- An entirely ordinary personal customer at onboarding: a current account, digital onboarding,
  electronic verification, income under £5,000 a month, nothing inconsistent on file.
""".strip(), """
Very little is open here, and pushing what is open to its worst defensible level moves the score by
about a tenth of a point. Both readings are Low.
""".strip(), """
1. **This is the right answer.** The lowest score in the back-test, on a customer who was, at
   onboarding, exactly what the file said. Nothing in a customer risk assessment could have predicted
   a fraud that had not happened yet against a scheme that did not yet exist.

2. **The failure is the refresh, not the rating.** The FCA's finding is that assessments were not kept
   up to date. This model produces a rating at onboarding and refreshes it on the cycle at methodology
   7.1, with event-driven triggers at 7.2. Trigger (d) — business activity on a personal account —
   exists in this project because of this notice.

3. **A back-test that only ever confirmed the model would be worthless.** This case is here because it
   is the one where the model is right and the interesting failure is somewhere else entirely.
""".strip())

print("wrote backtest README and %d case files" % len(CASES))
