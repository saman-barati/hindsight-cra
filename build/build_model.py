# -*- coding: utf-8 -*-
"""Build model/customer-risk-model.xlsx for Hindsight Step 3."""
import os
from paths import REPO
import sys, csv
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data import CATEGORIES, FACTORS
from labels import LABELS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter as L

CSV = REPO + "/data/synthetic-customers.csv"
OUT = REPO + "/model/customer-risk-model.xlsx"

FACTOR_ORDER = [f[0] for f in FACTORS]
CAT_OF = {f[0]: f[1] for f in FACTORS}
WIC = {f[0]: f[3] for f in FACTORS}
CATW = {c[0]: c[2] for c in CATEGORIES}
CATNAME = {c[0]: c[1] for c in CATEGORIES}

rows = list(csv.DictReader(open(CSV, encoding="utf-8")))
NR = len(rows)
R0, R1 = 2, 1 + NR                      # data rows

ARIAL = "Arial"
HDR = PatternFill("solid", fgColor="14454E")
SUB = PatternFill("solid", fgColor="E4EAEA")
IN_ = PatternFill("solid", fgColor="FFF2CC")
OKF = PatternFill("solid", fgColor="E2EFDA")
HIF = PatternFill("solid", fgColor="F6E2E2")
H = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
B = Font(name=ARIAL, size=10, bold=True)
N = Font(name=ARIAL, size=10)
MU = Font(name=ARIAL, size=9, color="5C6E72")
TITLE = Font(name=ARIAL, size=14, bold=True, color="14454E")
CTR = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
BOX = Border(bottom=Side(style="thin", color="C9D4D4"))

wb = Workbook()

def head(ws, r, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(r, i, h)
        c.font, c.fill, c.alignment = H, HDR, Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[L(i)].width = w
    ws.row_dimensions[r].height = 30

def table(ws, name, r0, r1, c1):
    t = Table(displayName=name, ref="A%d:%s%d" % (r0, L(c1), r1))
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight8", showRowStripes=True)
    ws.add_table(t)

# ------------------------------------------------------------------ Overview
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 106
lines = [
 (TITLE, "Customer risk scoring model"),
 (B, "Northgate Bank UK Limited (fictional)  |  HND-CRA-004  |  version 0.4 draft  |  September 2026"),
 (N, ""),
 (B, "What this workbook does"),
 (N, "Step 3 of the Hindsight project. It takes the %d synthetic customers in data/synthetic-customers.csv, "
     "converts each recorded level into a score using the Step 2 risk factor library, applies the category "
     "weights, produces an overall score and a band, and then applies the mandatory escalators from section 5.3 "
     "of the methodology." % NR),
 (N, ""),
 (B, "Sheets"),
 (N, "Mapping    -  the 98 level definitions from the Step 2 library, each with its score. This is the lookup table."),
 (N, "Weights    -  the 20 factors with their weight inside their category and their effective weight across the model."),
 (N, "Customers  -  the imported population. One row per customer, one column per factor, holding the level recorded at onboarding."),
 (N, "Scoring    -  the engine. Each factor label is converted to a score, weighted, summed, banded, then escalated."),
 (N, "Dashboard  -  what the population looks like once scored, and how much each category is actually doing."),
 (N, "Backtest   -  Step 4. Six FCA enforcement cases, each rebuilt as it looked at onboarding and scored two or three ways."),
 (N, "Validation -  Step 5. Weight sensitivity, three alternative aggregation rules, and the 2026 EDD test."),
 (N, "Checks     -  nineteen integrity tests. All must read OK."),
 (N, ""),
 (B, "How a rating is produced, column by column on the Scoring sheet"),
 (N, "C to V   score for each of the 20 factors, looked up from Mapping on the key 'factor|level'"),
 (N, "W        weighted score = SUMPRODUCT of the 20 scores and the 20 effective weights"),
 (N, "X        the same figure rounded to two decimal places, as required by methodology 5.2"),
 (N, "Y        band from the arithmetic alone: Low 1.00 to 2.00, Medium 2.01 to 3.49, High 3.50 to 5.00"),
 (N, "Z        whether any mandatory escalator in methodology 5.3 applies"),
 (N, "AA       the final customer risk rating. High if Z is Yes, otherwise the band in Y"),
 (N, "AB to AF the five category scores, kept so that the Dashboard can show which categories discriminate"),
 (N, ""),
 (B, "Refreshing the population with Power Query"),
 (N, "The Customers sheet holds the result of the import. To rebuild it from the CSV: Data > Get Data > From File > "
     "From Text/CSV, select data/synthetic-customers.csv, set the delimiter to comma and the encoding to UTF-8, "
     "promote the first row to headers, set every factor column to Text and onboarded to Date, then Close & Load To "
     "an existing worksheet at Customers!$A$1. The exact M code is in docs/03-model-build.md."),
 (N, ""),
 (B, "Which cells to edit"),
 (N, "None of the cells in this workbook are inputs. The weights live in the Step 2 library and are copied here; "
     "the scores come from the CSV. To change a weight, change it in model/risk-factor-library.xlsx first and copy "
     "it across, so that the two files cannot drift apart."),
 (N, ""),
 (MU, "Synthetic data only. Northgate Bank UK Limited does not exist, no real customer data appears in this workbook, "
      "and this is a self-directed learning project rather than professional compliance advice."),
]
r = 2
for font, text in lines:
    c = ws.cell(r, 2, text); c.font, c.alignment = font, WRAP
    if len(text) > 106:
        ws.row_dimensions[r].height = 14 * (len(text) // 106 + 1)
    r += 1
ws.freeze_panes = "A3"

# ------------------------------------------------------------------ Mapping
ws = wb.create_sheet("Mapping")
head(ws, 1, ["Factor", "Level recorded", "Score", "Lookup key"], [10, 44, 9, 48])
mr = 2
for f in FACTOR_ORDER:
    for score in sorted(LABELS[f]):
        ws.cell(mr, 1, f).font = B
        ws.cell(mr, 2, LABELS[f][score]).font = N
        c = ws.cell(mr, 3, score); c.font, c.alignment = N, CTR
        c = ws.cell(mr, 4, '=$A%d&"|"&$B%d' % (mr, mr)); c.font = N
        mr += 1
MR1 = mr - 1
ws.freeze_panes = "A2"
table(ws, "tblMapping", 1, MR1, 4)

# ------------------------------------------------------------------ Weights
ws = wb.create_sheet("Weights")
head(ws, 1, ["Factor", "Category", "Effective weight", "Weight in category", "", "", "Category", "Category weight"],
     [10, 12, 16, 17, 3, 3, 20, 16])
for i, f in enumerate(FACTOR_ORDER):
    r = 2 + i
    ws.cell(r, 1, f).font = B
    c = ws.cell(r, 2, CAT_OF[f]); c.font, c.alignment = N, CTR
    c = ws.cell(r, 3, round(CATW[CAT_OF[f]] * WIC[f], 10)); c.font, c.number_format, c.alignment = N, "0.00%", CTR
    c = ws.cell(r, 4, WIC[f]); c.font, c.number_format, c.alignment = N, "0%", CTR
WR1 = 1 + len(FACTOR_ORDER)
for j, (code, name, w) in enumerate(CATEGORIES):
    r = 2 + j
    ws.cell(r, 7, name).font = N
    c = ws.cell(r, 8, w); c.font, c.number_format, c.alignment = N, "0%", CTR
    ws.cell(r, 7).border = BOX; ws.cell(r, 8).border = BOX
c = ws.cell(WR1 + 1, 3, "=SUM($C$2:$C$%d)" % WR1)
c.font, c.number_format, c.alignment, c.fill = B, "0.00%", CTR, SUB
ws.cell(WR1 + 1, 2, "Total").font = B
ws.cell(WR1 + 3, 1, "Copied from model/risk-factor-library.xlsx v0.3. Check 4 below confirms the total is still 100%.").font = MU

# Horizontal copy of the two weight columns. The Scoring sheet holds one factor per
# column, so SUMPRODUCT needs the weights laid out across a row rather than down a
# column. These are formulas pointing at the table above, so there is still one source
# of truth for every weight.
HROW = WR1 + 5                       # 26: factor codes
ws.cell(HROW, 1, "Horizontal copy used by the Scoring sheet").font = B
ws.cell(HROW + 1, 1, "Effective weight").font = N
ws.cell(HROW + 2, 1, "Weight in category").font = N
for i in range(len(FACTOR_ORDER)):
    col = 3 + i
    c = ws.cell(HROW, col, "=$A%d" % (2 + i)); c.font, c.alignment = B, CTR
    c = ws.cell(HROW + 1, col, "=$C%d" % (2 + i)); c.font, c.number_format, c.alignment = N, "0.00%", CTR
    c = ws.cell(HROW + 2, col, "=$D%d" % (2 + i)); c.font, c.number_format, c.alignment = N, "0%", CTR
W_EFF = "Weights!$C$%d:$V$%d" % (HROW + 1, HROW + 1)
W_CAT_ROW = HROW + 2
ws.freeze_panes = "A2"

# ------------------------------------------------------------------ Customers
ws = wb.create_sheet("Customers")
cols = ["customer_id", "segment"] + FACTOR_ORDER + ["sar_last_12m", "onboarded"]
head(ws, 1, cols, [12, 16] + [22] * 20 + [13, 12])
for i, row in enumerate(rows):
    r = 2 + i
    for j, col in enumerate(cols, 1):
        c = ws.cell(r, j, row[col])
        c.font = N
        if j > 2:
            c.alignment = TOP
ws.freeze_panes = "C2"
table(ws, "tblCustomers", 1, R1, len(cols))

# ------------------------------------------------------------------ Scoring
ws = wb.create_sheet("Scoring")
sc_heads = ["customer_id", "segment"] + FACTOR_ORDER + [
    "Weighted score", "Rounded", "Band", "Escalator", "Final rating",
    "Customer", "Geography", "Product", "Channel", "Activity"]
head(ws, 1, sc_heads, [12, 16] + [7] * 20 + [14, 10, 11, 11, 12] + [11] * 5)
F_C, F_V = 3, 22          # factor score columns
W_, X_, Y_, Z_, AA_ = 23, 24, 25, 26, 27
AB_ = 28
CAT_RANGES = [("C", 3, 7, 2, 6), ("G", 8, 10, 7, 9), ("P", 11, 14, 10, 13),
              ("D", 15, 17, 14, 16), ("A", 18, 22, 17, 21)]
for i in range(NR):
    r = 2 + i
    ws.cell(r, 1, "=Customers!$A%d" % r).font = N
    ws.cell(r, 2, "=Customers!$B%d" % r).font = N
    for j in range(F_C, F_V + 1):
        c = ws.cell(r, j, '=INDEX(Mapping!$C$2:$C$%d,MATCH(%s$1&"|"&Customers!%s%d,Mapping!$D$2:$D$%d,0))'
                    % (MR1, L(j), L(j), r, MR1))
        c.font, c.alignment = N, CTR
    c = ws.cell(r, W_, "=SUMPRODUCT($C%d:$V%d,%s)" % (r, r, W_EFF))
    c.font, c.number_format, c.alignment = N, "0.0000", CTR
    c = ws.cell(r, X_, "=ROUND($W%d,2)" % r); c.font, c.number_format, c.alignment = N, "0.00", CTR
    c = ws.cell(r, Y_, '=IF($X%d<=2,"Low",IF($X%d<=3.49,"Medium","High"))' % (r, r)); c.font, c.alignment = N, CTR
    c = ws.cell(r, Z_, '=IF(OR($F%d>=3,$E%d=5,$D%d=5,$C%d=5,$G%d=5,$H%d=5,$I%d=5,$J%d=5,Customers!$W%d="Yes"),"Yes","No")'
                % (r, r, r, r, r, r, r, r, r))
    c.font, c.alignment = N, CTR
    c = ws.cell(r, AA_, '=IF($Z%d="Yes","High",$Y%d)' % (r, r)); c.font, c.alignment = B, CTR
    for k, (code, c0, c1, w0, w1) in enumerate(CAT_RANGES):
        c = ws.cell(r, AB_ + k, "=SUMPRODUCT($%s%d:$%s%d,Weights!$%s$%d:$%s$%d)"
                    % (L(c0), r, L(c1), r, L(c0), W_CAT_ROW, L(c1), W_CAT_ROW))
        c.font, c.number_format, c.alignment = N, "0.00", CTR
ws.freeze_panes = "C2"

# ------------------------------------------------------------------ Dashboard
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False
for col, w in [("A", 3), ("B", 40), ("C", 12), ("D", 12), ("E", 12), ("F", 12), ("G", 12), ("H", 40)]:
    ws.column_dimensions[col].width = w
ws.cell(1, 2, "Population of %d customers, scored" % NR).font = TITLE
ws.cell(2, 2, "Everything on this sheet is calculated from the Scoring sheet. Nothing is typed in.").font = MU

def block(r, title):
    c = ws.cell(r, 2, title); c.font = B; return r + 1

def hdr(r, labels, start=2):
    for i, t in enumerate(labels):
        c = ws.cell(r, start + i, t); c.font, c.fill, c.alignment = H, HDR, CTR
    return r + 1

SC = "Scoring!"
r = block(4, "Final customer risk rating")
r = hdr(r, ["Rating", "Customers", "Share"])
rate_start = r
for band in ["Low", "Medium", "High"]:
    ws.cell(r, 2, band).font = N
    c = ws.cell(r, 3, '=COUNTIF(%s$AA$2:$AA$%d,$B%d)' % (SC, R1, r)); c.font, c.alignment = N, CTR
    c = ws.cell(r, 4, "=$C%d/%d" % (r, NR)); c.font, c.number_format, c.alignment = N, "0.0%", CTR
    if band == "High":
        ws.cell(r, 2).fill = HIF
    r += 1
ws.cell(r, 2, "Total").font = B
c = ws.cell(r, 3, "=SUM($C$%d:$C$%d)" % (rate_start, r - 1)); c.font, c.alignment, c.fill = B, CTR, SUB
rate_end = r - 1
r += 2

r = block(r, "Band from the arithmetic alone, before escalators")
r = hdr(r, ["Band", "Customers", "Share"])
band_start = r
for band in ["Low", "Medium", "High"]:
    ws.cell(r, 2, band).font = N
    c = ws.cell(r, 3, '=COUNTIF(%s$Y$2:$Y$%d,$B%d)' % (SC, R1, r)); c.font, c.alignment = N, CTR
    c = ws.cell(r, 4, "=$C%d/%d" % (r, NR)); c.font, c.number_format, c.alignment = N, "0.0%", CTR
    r += 1
ws.cell(r, 2, "Lifted to High by an escalator").font = B
c = ws.cell(r, 3, '=COUNTIFS(%s$Y$2:$Y$%d,"<>High",%s$AA$2:$AA$%d,"High")' % (SC, R1, SC, R1))
c.font, c.alignment, c.fill = B, CTR, OKF
r += 3

r = block(r, "Why customers were escalated")
r = hdr(r, ["Escalator", "Customers"])
esc = [
 ("5.3(a)  PEP, family member or close associate", '=COUNTIF(%s$F$2:$F$%d,">=3")' % (SC, R1)),
 ("5.3(b)  Call for action jurisdiction", "=SUMPRODUCT(--((%s$H$2:$H$%d=5)+(%s$I$2:$I$%d=5)+(%s$J$2:$J$%d=5)>0))" % (SC, R1, SC, R1, SC, R1)),
 ("5.3(c)  MSB, TCSP or high-value dealer", "=COUNTIF(%s$E$2:$E$%d,5)" % (SC, R1)),
 ("5.3(d)  Opaque ownership structure", "=SUMPRODUCT(--((%s$D$2:$D$%d=5)+(%s$C$2:$C$%d=5)>0))" % (SC, R1, SC, R1)),
 ("5.3(e)  Adverse media, financial crime", "=COUNTIF(%s$G$2:$G$%d,5)" % (SC, R1)),
 ("5.3(f)  SAR in the preceding 12 months", '=COUNTIF(Customers!$W$2:$W$%d,"Yes")' % R1),
 ("Any escalator (customers, not reasons)", '=COUNTIF(%s$Z$2:$Z$%d,"Yes")' % (SC, R1)),
 ("Escalated by 5.3(f) and by nothing else",
  '=SUMPRODUCT((Customers!$W$2:$W$%d="Yes")*--((%s$F$2:$F$%d>=3)+(%s$E$2:$E$%d=5)+(%s$D$2:$D$%d=5)'
  '+(%s$C$2:$C$%d=5)+(%s$G$2:$G$%d=5)+(%s$H$2:$H$%d=5)+(%s$I$2:$I$%d=5)+(%s$J$2:$J$%d=5)=0))'
  % (R1, SC, R1, SC, R1, SC, R1, SC, R1, SC, R1, SC, R1, SC, R1, SC, R1)),
]
for label, formula in esc:
    ws.cell(r, 2, label).font = B if label.startswith("Any") else N
    c = ws.cell(r, 3, formula); c.font, c.alignment = (B if label.startswith("Any") else N), CTR
    r += 1
r += 2

r = block(r, "Rating by segment")
r = hdr(r, ["Segment", "Low", "Medium", "High", "Total"])
seg_start = r
for seg in ["Personal", "Sole trader", "Limited company"]:
    ws.cell(r, 2, seg).font = N
    for k, band in enumerate(["Low", "Medium", "High"]):
        c = ws.cell(r, 3 + k, '=COUNTIFS(%s$B$2:$B$%d,$B%d,%s$AA$2:$AA$%d,"%s")' % (SC, R1, r, SC, R1, band))
        c.font, c.alignment = N, CTR
    c = ws.cell(r, 6, "=SUM($C%d:$E%d)" % (r, r)); c.font, c.alignment = B, CTR
    r += 1
r += 2

r = block(r, "How much work each category is doing")
r = hdr(r, ["Category", "Mean", "Lowest", "Highest", "Std dev", "At the minimum"])
cat_start = r
for k, (code, c0, c1, w0, w1) in enumerate(CAT_RANGES):
    col = L(AB_ + k)
    ws.cell(r, 2, CATNAME[code]).font = N
    for i, fml in enumerate([
        "=AVERAGE(%s$%s$2:$%s$%d)" % (SC, col, col, R1),
        "=MIN(%s$%s$2:$%s$%d)" % (SC, col, col, R1),
        "=MAX(%s$%s$2:$%s$%d)" % (SC, col, col, R1),
        "=STDEV(%s$%s$2:$%s$%d)" % (SC, col, col, R1),
    ]):
        c = ws.cell(r, 3 + i, fml); c.font, c.number_format, c.alignment = N, "0.00", CTR
    c = ws.cell(r, 7, "=SUMPRODUCT(--(%s$%s$2:$%s$%d=MIN(%s$%s$2:$%s$%d)))/%d"
                % (SC, col, col, R1, SC, col, col, R1, NR))
    c.font, c.number_format, c.alignment = N, "0.0%", CTR
    r += 1
ws.cell(r, 2, "Standard deviation is the measure that matters. A category whose score barely varies across the "
              "population is not separating customers, whatever weight it carries.").font = MU
ws.cell(r, 2).alignment = WRAP
ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=7)
r += 3

r = block(r, "Overall score")
r = hdr(r, ["Statistic", "Value"])
for label, fml in [("Mean", "=AVERAGE(%s$W$2:$W$%d)" % (SC, R1)),
                   ("Lowest", "=MIN(%s$W$2:$W$%d)" % (SC, R1)),
                   ("Highest", "=MAX(%s$W$2:$W$%d)" % (SC, R1)),
                   ("Standard deviation", "=STDEV(%s$W$2:$W$%d)" % (SC, R1)),
                   ("Within 0.10 of the Low/Medium boundary",
                    '=COUNTIFS(%s$X$2:$X$%d,">=1.91",%s$X$2:$X$%d,"<=2.10")' % (SC, R1, SC, R1))]:
    ws.cell(r, 2, label).font = N
    c = ws.cell(r, 3, fml); c.font, c.alignment = N, CTR
    c.number_format = "0.00" if label != "Within 0.10 of the Low/Medium boundary" else "General"
    r += 1
r += 2

r = block(r, "Distribution of the overall score")
r = hdr(r, ["From", "To", "Customers"])
hist_start = r
lo = 1.0
while lo < 5.0 - 1e-9:
    hi = round(lo + 0.25, 2)
    c = ws.cell(r, 2, lo); c.font, c.number_format, c.alignment = N, "0.00", CTR
    c = ws.cell(r, 3, hi); c.font, c.number_format, c.alignment = N, "0.00", CTR
    c = ws.cell(r, 4, '=COUNTIFS(%s$W$2:$W$%d,">=%.2f",%s$W$2:$W$%d,"<%.2f")' % (SC, R1, lo, SC, R1, hi))
    c.font, c.alignment = N, CTR
    r += 1
    lo = hi
hist_end = r - 1
ws.freeze_panes = "A4"

ch = BarChart(); ch.type = "col"; ch.style = 10
ch.title = "Final customer risk rating"
ch.y_axis.title = "Customers"; ch.x_axis.title = None
ch.add_data(Reference(ws, min_col=3, min_row=rate_start - 1, max_row=rate_end), titles_from_data=True)
ch.set_categories(Reference(ws, min_col=2, min_row=rate_start, max_row=rate_end))
ch.height, ch.width, ch.legend = 6.5, 11, None
ws.add_chart(ch, "J4")

ch2 = BarChart(); ch2.type = "col"; ch2.style = 10
ch2.title = "Overall score distribution"
ch2.y_axis.title = "Customers"; ch2.x_axis.title = "Score, lower bound"
ch2.add_data(Reference(ws, min_col=4, min_row=hist_start - 1, max_row=hist_end), titles_from_data=True)
ch2.set_categories(Reference(ws, min_col=2, min_row=hist_start, max_row=hist_end))
ch2.height, ch2.width, ch2.legend = 7.5, 15, None
ch2.gapWidth = 20
ws.add_chart(ch2, "J20")

# ------------------------------------------------------------------ Checks
ws = wb.create_sheet("Checks")
ws.sheet_view.showGridLines = False
ws.cell(1, 1, "Integrity checks").font = TITLE
ws.cell(2, 1, "Every row must read OK. These recalculate whenever the population or a weight changes.").font = MU
head(ws, 4, ["#", "What is tested", "Result", "Expected", "Status"], [5, 76, 12, 12, 12])
CH = [
 ("The population contains %d customers" % NR, "=COUNTA(%s$A$2:$A$%d)" % (SC, R1), NR),
 ("Every recorded level resolves to a score in the library", "=SUMPRODUCT(--ISNA(%s$C$2:$V$%d))" % (SC, R1), 0),
 ("No factor score falls outside the range 1 to 5", "=SUMPRODUCT(--((%s$C$2:$V$%d<1)+(%s$C$2:$V$%d>5)>0))" % (SC, R1, SC, R1), 0),
 ("Effective weights still sum to 100%", "=ROUND(SUM(Weights!$C$2:$C$%d),6)" % WR1, 1),
 ("Weights inside the five categories sum to 100% each", "=ROUND(SUM(Weights!$D$2:$D$%d),6)" % WR1, 5),
 ("No overall score below 1.00", "=SUMPRODUCT(--(%s$W$2:$W$%d<1))" % (SC, R1), 0),
 ("No overall score above 5.00", "=SUMPRODUCT(--(%s$W$2:$W$%d>5))" % (SC, R1), 0),
 ("Where no escalator applies, the final rating equals the band",
  '=SUMPRODUCT(--(%s$Z$2:$Z$%d="No"),--(%s$AA$2:$AA$%d<>%s$Y$2:$Y$%d))' % (SC, R1, SC, R1, SC, R1), 0),
 ("Every escalated customer is rated High",
  '=SUMPRODUCT(--(%s$Z$2:$Z$%d="Yes"),--(%s$AA$2:$AA$%d<>"High"))' % (SC, R1, SC, R1), 0),
 ("No duplicate customer reference",
  "=SUMPRODUCT(--(COUNTIF(%s$A$2:$A$%d,%s$A$2:$A$%d)>1))" % (SC, R1, SC, R1), 0),
 ("The five category scores reconstruct the overall score",
  "=SUMPRODUCT(--(ROUND(%s$AB$2:$AB$%d*Weights!$H$2+%s$AC$2:$AC$%d*Weights!$H$3+%s$AD$2:$AD$%d*Weights!$H$4"
  "+%s$AE$2:$AE$%d*Weights!$H$5+%s$AF$2:$AF$%d*Weights!$H$6,6)<>ROUND(%s$W$2:$W$%d,6)))"
  % (SC, R1, SC, R1, SC, R1, SC, R1, SC, R1, SC, R1), 0),
 ("No duplicate lookup key in the mapping table",
  "=SUMPRODUCT(--(COUNTIF(Mapping!$D$2:$D$%d,Mapping!$D$2:$D$%d)>1))" % (MR1, MR1), 0),
]
for i, (label, fml, exp) in enumerate(CH):
    r = 5 + i
    c = ws.cell(r, 1, i + 1); c.font, c.alignment = N, CTR
    c = ws.cell(r, 2, label); c.font, c.alignment = N, WRAP
    c = ws.cell(r, 3, fml); c.font, c.alignment = N, CTR
    c = ws.cell(r, 4, exp); c.font, c.alignment = N, CTR
    c = ws.cell(r, 5, '=IF($C%d=$D%d,"OK","CHECK")' % (r, r)); c.font, c.alignment, c.fill = B, CTR, OKF
    for k in range(1, 6):
        ws.cell(r, k).border = BOX
r = 5 + len(CH) + 1
ws.cell(r, 2, "Overall status").font = B
c = ws.cell(r, 5, '=IF(COUNTIF($E$5:$E$%d,"CHECK")=0,"ALL OK","FIX")' % (4 + len(CH)))
c.font, c.alignment, c.fill = B, CTR, OKF
ws.freeze_panes = "A5"

# ------------------------------------------------------------------ Step 5 columns
from validation_block import SCENARIOS, BAND_B_LOW, BAND_B_MED, SEV_MED, SEV_HIGH, KEYS, NAMES
wsS = wb["Scoring"]
SCN0 = 34                                  # AH : ten sensitivity scenario scores
SEVC = SCN0 + len(SCENARIOS)               # AR : count of factors scoring 4 or 5
RB, RC, RD, RDF = SEVC + 1, SEVC + 2, SEVC + 3, SEVC + 4
hdr = ([lbl for lbl, _, _ in SCENARIOS]
       + ["Factors at 4+", "Rule B band", "Rule C band", "Rule D band", "Rule D final"])
for i, t in enumerate(hdr):
    c = wsS.cell(1, SCN0 + i, t)
    c.font, c.fill = H, HDR
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wsS.column_dimensions[L(SCN0 + i)].width = 13
for i in range(NR):
    r = 2 + i
    for k in range(len(SCENARIOS)):
        c = wsS.cell(r, SCN0 + k, "=SUMPRODUCT($AB%d:$AF%d,Validation!$D$%d:$H$%d)" % (r, r, 7 + k, 7 + k))
        c.font, c.number_format, c.alignment = N, "0.00", CTR
    c = wsS.cell(r, SEVC, '=COUNTIF($C%d:$V%d,">=4")' % (r, r)); c.font, c.alignment = N, CTR
    c = wsS.cell(r, RB, '=IF($X%d<=%s,"Low",IF($X%d<=%s,"Medium","High"))'
                 % (r, BAND_B_LOW, r, BAND_B_MED)); c.font, c.alignment = N, CTR
    c = wsS.cell(r, RC, '=IF($%s%d=0,"Low",IF($%s%d<%d,"Medium","High"))'
                 % (L(SEVC), r, L(SEVC), r, SEV_HIGH)); c.font, c.alignment = N, CTR
    c = wsS.cell(r, RD, '=IF(OR($%s%d="High",$%s%d="High"),"High",'
                        'IF(OR($%s%d="Medium",$%s%d="Medium"),"Medium","Low"))'
                 % (L(RB), r, L(RC), r, L(RB), r, L(RC), r)); c.font, c.alignment = N, CTR
    c = wsS.cell(r, RDF, '=IF($Z%d="Yes","High",$%s%d)' % (r, L(RD), r)); c.font, c.alignment = B, CTR

# ------------------------------------------------------------------ Validation
ws = wb.create_sheet("Validation")
ws.sheet_view.showGridLines = False
for col, wd in [("A", 3), ("B", 40), ("C", 26), ("D", 11), ("E", 11), ("F", 11), ("G", 11),
                ("H", 11), ("I", 15), ("J", 15)]:
    ws.column_dimensions[col].width = wd
ws.cell(1, 2, "Step 5: model validation").font = TITLE
c = ws.cell(2, 2, "Everything on this sheet is calculated from the Scoring sheet. The weights in columns D to H "
                  "are the only typed numbers, and they are the sensitivity scenarios themselves.")
c.font, c.alignment = MU, WRAP
ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=10)

SCn = "Scoring!"
r = 5
ws.cell(r, 2, "1. Weight sensitivity").font = B
r += 1
for i, t in enumerate(["Scenario", "Effect", "C", "G", "P", "D", "A",
                       "Band changes, current bands", "Band changes, rebanded"]):
    c = ws.cell(r, 2 + i, t); c.font, c.fill, c.alignment = H, HDR, CTR
SENS0 = r + 1
for k, (lbl, eff, wts) in enumerate(SCENARIOS):
    rr = SENS0 + k
    ws.cell(rr, 2, lbl).font = N
    ws.cell(rr, 3, eff).font = N
    for j, v in enumerate(wts):
        c = ws.cell(rr, 4 + j, v); c.font, c.number_format, c.alignment, c.fill = Font(name=ARIAL, size=10, color="0000FF"), "0%", CTR, IN_
    col = L(SCN0 + k)
    c = ws.cell(rr, 9, "=SUMPRODUCT(--((1+(ROUND(%s$%s$2:$%s$%d,2)>2)+(ROUND(%s$%s$2:$%s$%d,2)>3.49))"
                       "<>(1+(%s$X$2:$X$%d>2)+(%s$X$2:$X$%d>3.49))))"
                % (SCn, col, col, R1, SCn, col, col, R1, SCn, R1, SCn, R1))
    c.font, c.alignment = N, CTR
    c = ws.cell(rr, 10, "=SUMPRODUCT(--((1+(ROUND(%s$%s$2:$%s$%d,2)>%s)+(ROUND(%s$%s$2:$%s$%d,2)>%s))"
                        "<>(1+(%s$X$2:$X$%d>%s)+(%s$X$2:$X$%d>%s))))"
                % (SCn, col, col, R1, BAND_B_LOW, SCn, col, col, R1, BAND_B_MED,
                   SCn, R1, BAND_B_LOW, SCn, R1, BAND_B_MED))
    c.font, c.alignment = N, CTR
SENS1 = SENS0 + len(SCENARIOS) - 1
r = SENS1 + 2
c = ws.cell(r, 2, "Out of %d customers. Read the two right-hand columns together: the same weight change moves "
                  "far more customers once the band boundaries sit where the population actually is." % NR)
c.font, c.alignment = MU, WRAP
ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=10)

r += 3
ws.cell(r, 2, "2. Aggregation rules compared").font = B
r += 1
for i, t in enumerate(["Rule", "Definition", "Low", "Medium", "High", "High share"]):
    c = ws.cell(r, 2 + i, t); c.font, c.fill, c.alignment = H, HDR, CTR
RUL0 = r + 1
RULES = [
 ("A  current", "Weighted average, bands 1.00 / 2.01 / 3.50", "$AA"),
 ("B  rebanded", "Weighted average, bands 1.00 / %.2f / %.2f" % (BAND_B_LOW + 0.01, BAND_B_MED + 0.01), None),
 ("C  severity", "Count of factors scoring 4 or 5: 0 Low, 1-2 Medium, 3+ High", None),
 ("D  higher of B and C", "The worse of the two ratings above", None),
]
for k, (nm, dfn, col) in enumerate(RULES):
    rr = RUL0 + k
    ws.cell(rr, 2, nm).font = B
    c = ws.cell(rr, 3, dfn); c.font, c.alignment = N, WRAP
    for j, band in enumerate(["Low", "Medium", "High"]):
        if k == 0:
            f = '=COUNTIF(%s$AA$2:$AA$%d,"%s")' % (SCn, R1, band)
        elif k == 1:
            f = ('=SUMPRODUCT(--(IF(%s$Z$2:$Z$%d="Yes","High",%s$%s$2:$%s$%d)="%s"))'
                 % (SCn, R1, SCn, L(RB), L(RB), R1, band))
        elif k == 2:
            f = ('=SUMPRODUCT(--(IF(%s$Z$2:$Z$%d="Yes","High",%s$%s$2:$%s$%d)="%s"))'
                 % (SCn, R1, SCn, L(RC), L(RC), R1, band))
        else:
            f = '=COUNTIF(%s$%s$2:$%s$%d,"%s")' % (SCn, L(RDF), L(RDF), R1, band)
        c = ws.cell(rr, 4 + j, f); c.font, c.alignment = N, CTR
    c = ws.cell(rr, 7, "=$F%d/%d" % (rr, NR)); c.font, c.number_format, c.alignment = N, "0.0%", CTR
    if k == 3:
        ws.cell(rr, 7).fill = HIF
RUL1 = RUL0 + len(RULES) - 1
r = RUL1 + 2
c = ws.cell(r, 2, "The High column is the operational question. Rules C and D put roughly 1.7 times as many customers into "
                  "High as rule B does. The judgement that this is unaffordable is an assumption about cost, not a sourced "
                  "constraint: docs/04-model-validation.md section 4.2 says so, and names what would settle it.")
c.font, c.alignment = MU, WRAP
ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=10)

r += 3
ws.cell(r, 2, "3. The 2026 change to jurisdiction-based enhanced due diligence").font = B
r += 1
for i, t in enumerate(["Measure", "Customers", "Share of the book"]):
    c = ws.cell(r, 2 + i, t); c.font, c.fill, c.alignment = H, HDR, CTR
G4 = "((%s$H$2:$H$%d=4)+(%s$I$2:$I$%d=4)+(%s$J$2:$J$%d=4)>0)" % (SCn, R1, SCn, R1, SCn, R1)
G5 = "((%s$H$2:$H$%d=5)+(%s$I$2:$I$%d=5)+(%s$J$2:$J$%d=5)>0)" % (SCn, R1, SCn, R1, SCn, R1)
EDD = [
 ("Touching a FATF increased-monitoring jurisdiction", "=SUMPRODUCT(--%s)" % G4),
 ("Touching a FATF call-for-action jurisdiction", "=SUMPRODUCT(--%s)" % G5),
 ("Mandatory jurisdiction-based EDD before 2026", "=SUMPRODUCT(--(%s+%s>0))" % (G4, G5)),
 ("Mandatory jurisdiction-based EDD under the 2026 rule", "=SUMPRODUCT(--%s)" % G5),
 ("Customers losing automatic EDD", "=SUMPRODUCT(--(%s)*--(NOT(%s)))" % (G4, G5)),
 ("Of those, escalated for some other reason",
  '=SUMPRODUCT(--(%s),--(NOT(%s)),--(%s$Z$2:$Z$%d="Yes"))' % (G4, G5, SCn, R1)),
 ("Of those, rated Low once EDD is no longer automatic",
  '=SUMPRODUCT(--(%s),--(NOT(%s)),--(%s$AA$2:$AA$%d="Low"))' % (G4, G5, SCn, R1)),
]
EDD0 = r + 1
for k, (lbl, f) in enumerate(EDD):
    rr = EDD0 + k
    c = ws.cell(rr, 2, lbl); c.font, c.alignment = (B if k >= 4 else N), WRAP
    c = ws.cell(rr, 3, f); c.font, c.alignment = (B if k >= 4 else N), CTR
    c = ws.cell(rr, 4, "=$C%d/%d" % (rr, NR)); c.font, c.number_format, c.alignment = N, "0.0%", CTR
    if k >= 4:
        ws.cell(rr, 3).fill = HIF
ws.freeze_panes = "A5"

# ------------------------------------------------------------------ Backtest
from backtest_block import BACKTEST
ws = wb.create_sheet("Backtest")
NF = len(FACTOR_ORDER)
LAB0 = 3
SCO0 = LAB0 + NF
RES0 = SCO0 + NF
ws.cell(1, 1, "Step 4: enforcement cases, scored at onboarding").font = TITLE
c = ws.cell(2, 1, "Each case is reconstructed as it looked on the day the bank took it on, using only what that "
                  "bank knew or recorded then, and is scored twice: -lo is the reading most favourable to the "
                  "customer that the published notice permits, -hi the least favourable reading a competent "
                  "analyst could have defended on the same information. Where the notice settles a factor, both "
                  "rows carry the same level. FO-lo and FO-mid hold the entire file constant and change only the "
                  "C3 reading. Two cases are labelled -a and -b instead, because their pair is not a "
                  "favourable/adverse bracket: Stunt \u0026 Co is dated (the application file of 16 January 2015, "
                  "then the same file after the meeting of 27 January 2015), and Santander is conditional (the file "
                  "as recorded, then the same file with the business verified as a money service business, which is "
                  "the verification the FCA found was missing). Same Mapping, same Weights and same rules as the "
                  "400 synthetic customers.")
c.font, c.alignment = MU, WRAP
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
bt_heads = (["Case", "Enforcement case"]
            + ["%s level" % f for f in FACTOR_ORDER]
            + list(FACTOR_ORDER)
            + ["Weighted score", "Rounded", "Band", "Escalator", "Final rating"]
            + ["Customer", "Geography", "Product", "Channel", "Activity"])
head(ws, 3, bt_heads, [7, 46] + [26] * NF + [7] * NF + [14, 10, 11, 11, 12] + [11] * 5)
BT0 = 4
def sc_col(code):
    return L(SCO0 + FACTOR_ORDER.index(code))
for i, (ref, name, prof) in enumerate(BACKTEST):
    r = BT0 + i
    c = ws.cell(r, 1, ref); c.font, c.alignment = B, CTR
    c = ws.cell(r, 2, name); c.font = N
    for j, f in enumerate(FACTOR_ORDER):
        c = ws.cell(r, LAB0 + j, prof[f]); c.font, c.alignment = N, TOP
        c = ws.cell(r, SCO0 + j,
                    '=INDEX(Mapping!$C$2:$C$%d,MATCH("%s|"&%s%d,Mapping!$D$2:$D$%d,0))'
                    % (MR1, f, L(LAB0 + j), r, MR1))
        c.font, c.alignment = N, CTR
    c = ws.cell(r, RES0, "=SUMPRODUCT($%s%d:$%s%d,%s)" % (L(SCO0), r, L(SCO0 + NF - 1), r, W_EFF))
    c.font, c.number_format, c.alignment = N, "0.0000", CTR
    c = ws.cell(r, RES0 + 1, "=ROUND($%s%d,2)" % (L(RES0), r))
    c.font, c.number_format, c.alignment = N, "0.00", CTR
    c = ws.cell(r, RES0 + 2, '=IF($%s%d<=2,"Low",IF($%s%d<=3.49,"Medium","High"))'
                % (L(RES0 + 1), r, L(RES0 + 1), r)); c.font, c.alignment = N, CTR
    c = ws.cell(r, RES0 + 3, '=IF(OR($%s%d>=3,$%s%d=5,$%s%d=5,$%s%d=5,$%s%d=5,$%s%d=5,$%s%d=5,$%s%d=5),"Yes","No")'
                % (sc_col("C4"), r, sc_col("C3"), r, sc_col("C2"), r, sc_col("C1"), r,
                   sc_col("C5"), r, sc_col("G1"), r, sc_col("G2"), r, sc_col("G3"), r))
    c.font, c.alignment = N, CTR
    c = ws.cell(r, RES0 + 4, '=IF($%s%d="Yes","High",$%s%d)' % (L(RES0 + 3), r, L(RES0 + 2), r))
    c.font, c.alignment = B, CTR
    for k, (code, c0, c1, w0, w1) in enumerate(CAT_RANGES):
        s0 = SCO0 + (c0 - 3); s1 = SCO0 + (c1 - 3)
        c = ws.cell(r, RES0 + 5 + k, "=SUMPRODUCT($%s%d:$%s%d,Weights!$%s$%d:$%s$%d)"
                    % (L(s0), r, L(s1), r, L(c0), W_CAT_ROW, L(c1), W_CAT_ROW))
        c.font, c.number_format, c.alignment = N, "0.00", CTR
    # Step 5: the same case under the recommended package
    c = ws.cell(r, RES0 + 10, '=IF($%s%d<=%s,"Low",IF($%s%d<=%s,"Medium","High"))'
                % (L(RES0 + 1), r, BAND_B_LOW, L(RES0 + 1), r, BAND_B_MED))
    c.font, c.alignment = N, CTR
    c = ws.cell(r, RES0 + 11, '=IF(OR($%s%d="Yes",$%s%d=5),"High",$%s%d)'
                % (L(RES0 + 3), r, sc_col("A5"), r, L(RES0 + 10), r))
    c.font, c.alignment, c.fill = B, CTR, OKF
BT1 = BT0 + len(BACKTEST) - 1
for i, t in enumerate(["Rule B band", "Recommended final"]):
    c = ws.cell(3, RES0 + 10 + i, t)
    c.font, c.fill = H, HDR
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[L(RES0 + 10 + i)].width = 15
ws.cell(BT1 + 2, 2, "Reconstructions rated High at onboarding").font = B
c = ws.cell(BT1 + 2, 3, '=COUNTIF($%s$%d:$%s$%d,"High")' % (L(RES0 + 4), BT0, L(RES0 + 4), BT1))
c.font, c.alignment, c.fill = B, CTR, OKF
ws.cell(BT1 + 3, 2, "Reconstructions rated High by the arithmetic alone").font = B
c = ws.cell(BT1 + 3, 3, '=COUNTIF($%s$%d:$%s$%d,"High")' % (L(RES0 + 2), BT0, L(RES0 + 2), BT1))
c.font, c.alignment, c.fill = B, CTR, HIF
ws.freeze_panes = "C4"

wsc = wb["Checks"]
extra = [
 ("Every backtest level resolves to a score in the library",
  "=SUMPRODUCT(--ISNA(Backtest!$%s$%d:$%s$%d))" % (L(SCO0), BT0, L(SCO0 + NF - 1), BT1), 0),
 ("The backtest uses the same weights as the population",
  "=ROUND(SUMPRODUCT(Backtest!$%s$%d:$%s$%d,%s)-Backtest!$%s$%d,9)"
  % (L(SCO0), BT0, L(SCO0 + NF - 1), BT0, W_EFF, L(RES0), BT0), 0),
 ("The two Fowler Oldfield rows differ in exactly one factor score",
  "=SUMPRODUCT(--(Backtest!$%s$%d:$%s$%d<>Backtest!$%s$%d:$%s$%d))"
  % (L(SCO0), BT0, L(SCO0 + NF - 1), BT0, L(SCO0), BT0 + 1, L(SCO0 + NF - 1), BT0 + 1), 1),
 ("No -hi row scores below its -lo row on any factor",
  "=SUMPRODUCT(--(Backtest!$%s$%d:$%s$%d>Backtest!$%s$%d:$%s$%d))"
  % (L(SCO0), BT0, L(SCO0 + NF - 1), BT0, L(SCO0), BT0 + 2, L(SCO0 + NF - 1), BT0 + 2), 0),
]
base = 5 + len(CH)
for i, (label, fml, exp) in enumerate(extra):
    r = base + i
    c = wsc.cell(r, 1, len(CH) + i + 1); c.font, c.alignment = N, CTR
    c = wsc.cell(r, 2, label); c.font, c.alignment = N, WRAP
    c = wsc.cell(r, 3, fml); c.font, c.alignment = N, CTR
    c = wsc.cell(r, 4, exp); c.font, c.alignment = N, CTR
    c = wsc.cell(r, 5, '=IF($C%d=$D%d,"OK","CHECK")' % (r, r)); c.font, c.alignment, c.fill = B, CTR, OKF
    for k in range(1, 6):
        wsc.cell(r, k).border = BOX
wsc = wb["Checks"]
CH5 = [
 ("Every sensitivity scenario still sums to 100%",
  "=SUMPRODUCT(--(ROUND(MMULT(Validation!$D$%d:$H$%d,{1;1;1;1;1}),6)<>1))" % (SENS0, SENS1), 0),
 ("Rule D is never softer than rule B or rule C",
  '=SUMPRODUCT(--((%s$%s$2:$%s$%d="High")+(%s$%s$2:$%s$%d="High")>0),--(%s$%s$2:$%s$%d<>"High"))'
  % (SCn, L(RB), L(RB), R1, SCn, L(RC), L(RC), R1, SCn, L(RD), L(RD), R1), 0),
 ("Every customer has a factor-severity count between 0 and 20",
  "=SUMPRODUCT(--((%s$%s$2:$%s$%d<0)+(%s$%s$2:$%s$%d>20)>0))"
  % (SCn, L(SEVC), L(SEVC), R1, SCn, L(SEVC), L(SEVC), R1), 0),
]
base5 = 5 + len(CH) + len(extra)
for i, (label, fml, exp) in enumerate(CH5):
    rr = base5 + i
    c = wsc.cell(rr, 1, len(CH) + len(extra) + i + 1); c.font, c.alignment = N, CTR
    c = wsc.cell(rr, 2, label); c.font, c.alignment = N, WRAP
    c = wsc.cell(rr, 3, fml); c.font, c.alignment = N, CTR
    c = wsc.cell(rr, 4, exp); c.font, c.alignment = N, CTR
    c = wsc.cell(rr, 5, '=IF($C%d=$D%d,"OK","CHECK")' % (rr, rr)); c.font, c.alignment, c.fill = B, CTR, OKF
    for kk in range(1, 6):
        wsc.cell(rr, kk).border = BOX
TOT5 = len(CH) + len(extra) + len(CH5)
for rr in range(base5 + len(CH5), base5 + len(CH5) + 4):
    for kk in range(1, 6):
        wsc.cell(rr, kk).value = None
rr = 5 + TOT5 + 1
wsc.cell(rr, 2, "Overall status").font = B
c = wsc.cell(rr, 5, '=IF(COUNTIF($E$5:$E$%d,"CHECK")=0,"ALL OK","FIX")' % (4 + TOT5))
c.font, c.alignment, c.fill = B, CTR, OKF

for s in wb.worksheets:
    s.sheet_properties.tabColor = "14454E"
wb.save(OUT)
print("saved", OUT, "| customers:", NR, "| mapping rows:", MR1 - 1)
